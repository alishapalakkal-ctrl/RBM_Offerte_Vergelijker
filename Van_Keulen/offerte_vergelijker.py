#!/usr/bin/env python3
"""
Offerte Vergelijker - Compare Van Keulen furniture offers
Compares a PDF offer against Revit quantities and NETTO reference prices.

Usage:  python offerte_vergelijker.py
"""

import re
import sys
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List, Dict
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

try:
    import pdfplumber
except ImportError:
    print("pdfplumber not found. Install with: pip install pdfplumber")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("pandas not found. Install with: pip install pandas openpyxl")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)

try:
    from rapidfuzz import fuzz, process as rfprocess
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False

try:
    from PIL import Image, ImageDraw, ImageFont, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

def _make_jumbo_logo(width: int = 180, height: int = 60) -> "Image.Image":
    """Draw the Jumbo logo: yellow background, white JUMBO text with black outline."""
    from PIL import Image, ImageDraw, ImageFont
    img = Image.new("RGBA", (width, height), "#FDC400")
    draw = ImageDraw.Draw(img)

    font = None
    for name in ("arialbd.ttf", "Arial Bold.ttf", "impact.ttf", "verdanab.ttf"):
        try:
            font = ImageFont.truetype(name, int(height * 0.72))
            break
        except OSError:
            pass
    if font is None:
        font = ImageFont.load_default()

    text = "JUMBO"
    bbox = draw.textbbox((0, 0), text, font=font)
    x = (width - (bbox[2] - bbox[0])) // 2 - bbox[0]
    y = (height - (bbox[3] - bbox[1])) // 2 - bbox[1]

    for dx, dy in [(-2, -2), (2, -2), (-2, 2), (2, 2),
                   (0, -2), (0, 2), (-2, 0), (2, 0)]:
        draw.text((x + dx, y + dy), text, font=font, fill="#1A1A1A")
    draw.text((x, y), text, font=font, fill="white")
    return img


DEFAULT_FOLDER = Path(
    "C:/Users/a.palakkal/OneDrive - Retail Bouw Management BV/IB/Van Keulen/Poeldijk"
)

# ─── Custom Budget Groups ───────────────────────────────────────────────────────
# Each group batches PDF items matching a keyword and compares the combined total
# against a specific row in the Budget sheet of the IB file.
CUSTOM_GROUPS = [
    {
        "name": "Cross sell",
        "label": "Bovenschap marketkoeling → Crosselling bladen (Budget rij 1466)",
        "filter": lambda desc: "bovenschap marketkoeling" in desc.lower(),
        "budget_row": 1466,
    },
    {
        "name": "Draadschap",
        "label": "Draadschap items → Draadschap 1000x510mm (Budget rij 2612)",
        "filter": lambda desc: desc.lower().startswith("draadschap"),
        "budget_row": 2612,
    },
]

# ─── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class PdfItem:
    art_nr: str
    description: str
    quantity: float
    unit_price: float
    total: float
    section: str = ""
    manual_ref: str = ""
    page: int = 0


@dataclass
class NettoItem:
    art_nr_jumbo: str
    manual_nr: str
    art_nr_leverancier: str
    groep: str
    description: str
    netto_price: float
    unit: str = ""


@dataclass
class IBItem:
    nummer: str       # Budget row number e.g. D.01.14.226
    description: str  # Column H – Artikelnaam
    manual: str       # Column I
    unit: str         # Column J – Eenheid
    quantity: float   # Column K – Aantal
    price: float      # Column L – Prijs
    pdf_art_nr: str   # Column BA – PDF article number added by user


@dataclass
class MatchResult:
    pdf_item: Optional[PdfItem] = None
    netto_item: Optional[NettoItem] = None
    ib_items: List[IBItem] = field(default_factory=list)
    pdf_match_method: str = "unmatched"   # exact_artnr | manual_code | fuzzy | unmatched
    ib_match_method: str = "unmatched"    # exact_artnr | fuzzy | unmatched
    confidence: float = 0.0

    @property
    def ib_qty(self) -> Optional[float]:
        return sum(i.quantity for i in self.ib_items) if self.ib_items else None

    @property
    def qty_diff(self) -> Optional[float]:
        if self.pdf_item and self.ib_items:
            return self.pdf_item.quantity - self.ib_qty
        return None

    @property
    def price_diff(self) -> Optional[float]:
        if self.pdf_item and self.netto_item and self.netto_item.netto_price:
            return round(self.pdf_item.unit_price - self.netto_item.netto_price, 4)
        return None


# ─── PDF Parser ────────────────────────────────────────────────────────────────

# Matches the numeric tail of an item line: QTY PRICE,DD TOTAL,DD
# QTY may be in Dutch thousands format: 1.390 = 1390
_ITEM_END = re.compile(
    r'\s+((?:\d{1,3}\.)*\d+)\s+((?:\d{1,3}\.)*\d+,\d{2})\s+((?:\d{1,3}\.)*\d+,\d{2})\s*$'
)
_ART_NR = re.compile(r'^(\d{3,7})\s+(.*)')
_MANUAL = re.compile(r'^Manual\s+(\S+)', re.IGNORECASE)

_SKIP_STARTS = (
    'Offertenummer', 'Projectnummer', 'Ontvangende', 'Art.nr.', 'Art. nr.',
    'Op alle', 'Van Keulen', 'Postbus', 'www.', 'Pagina ',
    'Leverdatum', 'Uw contact', 'Planning', 'Aanvangsdatum',
    'EUR ', 'FSC ', 'PEFC ', 'ISO', 'Met vriendelijke',
    'Naam in blokletters', 'Handtekening', 'Totaalbedrag', 'Dit document',
    'Opdrachtgever', 'Jumbo Supermarkten', 'Mevr.', 'Dhr.', 'Nederland',
    'Offertedatum', 'Vervaldatum', 'Verkoper', 'Uw referentie',
)
# Dutch postal code line e.g. "5460 AA VEGHEL 2685 EB POELDIJK"
_POSTAL_RE = re.compile(r'^\d{4}\s+[A-Z]{2}\s+')

# Section header: short line, no digits, no special chars — e.g. "Systeem", "Brood"
_SECTION_RE = re.compile(r'^[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ /\-]{1,35}$')


def _dutch(s: str) -> float:
    """'10.698,50' → 10698.5"""
    return float(s.replace('.', '').replace(',', '.'))


def _skip(line: str) -> bool:
    if not line or len(line) <= 1:   # single chars from rotated "Offerte" watermark
        return True
    if _POSTAL_RE.match(line):       # Dutch postal code line
        return True
    for s in _SKIP_STARTS:
        if line.startswith(s):
            return True
    return False


def parse_pdf(path: str) -> List[PdfItem]:
    """Extract priced line items from a Van Keulen offer PDF."""
    items: List[PdfItem] = []
    section = ""
    manual_ref = ""
    art_nr = ""
    lines: List[str] = []
    page_nr = 0

    def flush():
        nonlocal art_nr, lines
        if not art_nr:
            return
        text = " ".join(lines).strip()
        m = _ITEM_END.search(text)
        if m:
            desc = text[: m.start()].strip()
            qty = _dutch(m.group(1))   # handles "1.390" (Dutch thousands) → 1390.0
            price = _dutch(m.group(2))
            total = _dutch(m.group(3))
            if price > 0:
                items.append(PdfItem(
                    art_nr=art_nr, description=desc,
                    quantity=qty, unit_price=price, total=total,
                    section=section, manual_ref=manual_ref, page=page_nr,
                ))
        art_nr = ""
        lines = []

    with pdfplumber.open(path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            if page_num == 1:        # cover page — no items
                continue
            page_nr = page_num
            text = page.extract_text()
            if not text:
                continue

            for raw in text.split('\n'):
                line = raw.strip()
                if _skip(line):
                    continue

                m = _MANUAL.match(line)
                if m:
                    flush()
                    manual_ref = m.group(1)
                    continue

                # Skip any comment/note line starting with [ (even multi-line ones)
                if line.startswith('['):
                    continue

                m = _ART_NR.match(line)
                if m:
                    # If we're already inside an item whose text has no price yet,
                    # this digit-starting line is a dimension continuation,
                    # e.g. "2200 hoog, 1000 breed," inside a multi-line description.
                    if art_nr and not _ITEM_END.search(" ".join(lines)):
                        lines.append(line)
                        # Eager completion: flush as soon as price pattern is found
                        if _ITEM_END.search(" ".join(lines)):
                            flush()
                    else:
                        flush()
                        art_nr = m.group(1)
                        lines = [m.group(2)]
                    continue

                if art_nr:
                    lines.append(line)
                    # Eager completion: flush as soon as price pattern is found
                    if _ITEM_END.search(" ".join(lines)):
                        flush()
                else:
                    # potential section header (short, letters only)
                    if line and _SECTION_RE.match(line):
                        section = line

        flush()

    return items


# ─── Excel Readers ─────────────────────────────────────────────────────────────

def read_netto(path: str) -> List[NettoItem]:
    """
    Read the NETTO prijslijst Excel file.
    Row 0 = blank, Row 1 = header → data from row 2 onward.
    Columns:
      0=Art.nr.Jumbo  1=Manual nr  2=Art.nr.leverancier  3=Groep
      4=Omschrijving  5=Netto factuurprijs  11=Eenheid  16=Netto netto prijs
    """
    df = pd.read_excel(path, sheet_name=0, header=None)
    items: List[NettoItem] = []
    for _, row in df.iloc[2:].iterrows():
        def cell(i): return str(row.iloc[i]).strip() if pd.notna(row.iloc[i]) else ""

        art_nr_jumbo = cell(0)
        if not art_nr_jumbo or art_nr_jumbo in ('nan', 'None'):
            continue

        manual_nr    = cell(1)
        art_nr_lev   = cell(2)
        groep        = cell(3)
        description  = cell(4)
        unit         = cell(11)

        # Unusable art_nr values
        if art_nr_lev in ('nan', 'None', 'Geen artikelno.', 'Niet leverbaar'):
            art_nr_lev = ""

        # Price: prefer col 16 (netto netto prijs), fall back to col 5
        price = 0.0
        for col_idx in (16, 5):
            val = row.iloc[col_idx]
            if pd.notna(val):
                try:
                    price = float(val)
                    break
                except (ValueError, TypeError):
                    pass

        items.append(NettoItem(
            art_nr_jumbo=art_nr_jumbo, manual_nr=manual_nr,
            art_nr_leverancier=art_nr_lev, groep=groep,
            description=description, netto_price=price, unit=unit,
        ))
    return items


def read_ib_items(path: str, row_start: int = 1440, row_end: int = 2763) -> List[IBItem]:
    """
    Read Budget sheet rows row_start..row_end from the JUMBO IB xlsm file.
    Columns (0-indexed):
      5  = level marker ("2" = actual item row)
      6  = Nummer  (D.01.xx.xx)
      7  = Artikelnaam / description   (Excel column H)
      8  = Manual                      (Excel column I)
      9  = Eenheid / unit              (Excel column J)
      10 = Aantal / quantity           (Excel column K)
      11 = Prijs / price               (Excel column L)
      52 = PDF art_nr added by user    (Excel column BA)
    """
    df = pd.read_excel(path, sheet_name="Budget", header=None, engine="openpyxl")
    items: List[IBItem] = []
    for idx in range(row_start - 1, min(row_end, len(df))):
        row = df.iloc[idx]
        def cell(i):
            return str(row.iloc[i]).strip() if i < len(row) and pd.notna(row.iloc[i]) else ""
        # only level-2 rows are actual items
        if cell(5) != "2":
            continue
        description = cell(7)
        if not description or description in ("nan", "NaT", "0"):
            continue
        try:    quantity = float(row.iloc[10])
        except: quantity = 0.0
        try:    price    = float(row.iloc[11])
        except: price    = 0.0
        pdf_art_nr = cell(52) if len(row) > 52 else ""
        if pdf_art_nr in ("nan", "NaT"):
            pdf_art_nr = ""
        items.append(IBItem(
            nummer=cell(6), description=description,
            manual=cell(8), unit=cell(9),
            quantity=quantity, price=price,
            pdf_art_nr=pdf_art_nr,
        ))
    return items


# ─── Budget Reader ─────────────────────────────────────────────────────────────

def read_budget_row(path: str, row_num: int) -> dict:
    """Read one row from the Budget sheet of the IB xlsm file (1-based row number)."""
    df = pd.read_excel(path, sheet_name="Budget", header=None, engine="openpyxl")
    row = df.iloc[row_num - 1]
    def cell(i): return str(row.iloc[i]).strip() if pd.notna(row.iloc[i]) else ""
    try:
        aantal = float(row.iloc[10])
    except (ValueError, TypeError):
        aantal = 0.0
    try:
        prijs = float(row.iloc[11])
    except (ValueError, TypeError):
        prijs = 0.0
    return {
        "row_num": row_num,
        "nummer": cell(6),
        "artikelnaam": cell(7),
        "manual": cell(8),
        "eenheid": cell(9),
        "aantal": aantal,
        "prijs": prijs,
        "totaal": round(aantal * prijs, 2),
    }


def build_budget_groups(pdf_items: List[PdfItem], budget_path: str) -> list:
    """Match each CUSTOM_GROUP definition against PDF items and load its Budget reference row."""
    groups = []
    for cfg in CUSTOM_GROUPS:
        matched = [p for p in pdf_items if cfg["filter"](p.description)]
        budget_row = read_budget_row(budget_path, cfg["budget_row"])
        pdf_total = round(sum(p.total for p in matched), 2)
        diff_eur  = round(pdf_total - budget_row["totaal"], 2)
        diff_pct  = round(diff_eur / budget_row["totaal"] * 100, 1) if budget_row["totaal"] else None
        groups.append({
            "name":       cfg["name"],
            "label":      cfg["label"],
            "pdf_items":  matched,
            "budget_row": budget_row,
            "pdf_total":  pdf_total,
            "diff_eur":   diff_eur,
            "diff_pct":   diff_pct,
        })
    return groups


# ─── Matcher ───────────────────────────────────────────────────────────────────

def _norm(code: str) -> str:
    """Normalise a manual code for dict lookup."""
    return code.strip().upper()


def build_matches(
    pdf_items: List[PdfItem],
    netto_items: List[NettoItem],
    ib_items: List[IBItem],
    fuzzy_threshold: int = 70,
) -> List[MatchResult]:

    # ── NETTO lookup dicts ─────────────────────────────────────────────────
    netto_by_artnr: Dict[str, NettoItem] = {}
    netto_by_manual: Dict[str, List[NettoItem]] = {}
    for ni in netto_items:
        if ni.art_nr_leverancier:
            netto_by_artnr[ni.art_nr_leverancier] = ni
        mn = _norm(ni.manual_nr)
        if mn:
            netto_by_manual.setdefault(mn, []).append(ni)

    # ── IB lookup: col BA (pdf_art_nr) for exact match ────────────────────
    ib_by_pdf_artnr: Dict[str, IBItem] = {
        ii.pdf_art_nr: ii for ii in ib_items if ii.pdf_art_nr
    }
    ib_descs = [ii.description for ii in ib_items] if HAS_RAPIDFUZZ else []

    # Fuzzy description corpus for NETTO (built once)
    netto_descs = [n.description for n in netto_items] if HAS_RAPIDFUZZ else []

    results: List[MatchResult] = []

    for pi in pdf_items:
        mr = MatchResult(pdf_item=pi)

        # 1. Exact art.nr. → NETTO
        ni = netto_by_artnr.get(pi.art_nr)
        if ni:
            mr.netto_item = ni
            mr.pdf_match_method = "exact_artnr"
            mr.confidence = 1.0

        # 2. Manual ref header → NETTO
        if not mr.netto_item and pi.manual_ref:
            mn = _norm(pi.manual_ref)
            candidates = netto_by_manual.get(mn, [])
            if candidates:
                if HAS_RAPIDFUZZ and len(candidates) > 1:
                    descs = [c.description for c in candidates]
                    best = rfprocess.extractOne(
                        pi.description, descs, scorer=fuzz.token_set_ratio
                    )
                    mr.netto_item = candidates[descs.index(best[0])] if best else candidates[0]
                else:
                    mr.netto_item = candidates[0]
                mr.pdf_match_method = "manual_code"
                mr.confidence = 0.85

        # 3. Fuzzy description → NETTO
        if not mr.netto_item and HAS_RAPIDFUZZ:
            best = rfprocess.extractOne(
                pi.description, netto_descs,
                scorer=fuzz.token_set_ratio,
                score_cutoff=fuzzy_threshold,
            )
            if best:
                mr.netto_item = netto_items[netto_descs.index(best[0])]
                mr.pdf_match_method = "fuzzy"
                mr.confidence = best[1] / 100.0

        # ── IB matching: col BA exact first, then fuzzy description ───────
        ii = ib_by_pdf_artnr.get(pi.art_nr)
        if ii:
            mr.ib_items = [ii]
            mr.ib_match_method = "exact_artnr"
        elif HAS_RAPIDFUZZ and ib_descs:
            best = rfprocess.extractOne(
                pi.description, ib_descs,
                scorer=fuzz.token_set_ratio,
                score_cutoff=fuzzy_threshold,
            )
            if best:
                mr.ib_items = [ib_items[ib_descs.index(best[0])]]
                mr.ib_match_method = "fuzzy"

        results.append(mr)

    return results


# ─── Excel Exporter ────────────────────────────────────────────────────────────

_FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_FILL_ORANGE = PatternFill("solid", fgColor="FFCC00")
_FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
_FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_FILL_HEADER = PatternFill("solid", fgColor="4472C4")
_FONT_HEADER = Font(bold=True, color="FFFFFF")
_FONT_BOLD   = Font(bold=True)

_HEADERS = [
    "Art.nr. (PDF)", "Omschrijving (PDF)", "Sectie (PDF)", "Manual Ref (PDF)",
    "Match methode", "Confidence",
    "PDF Aantal", "IB Aantal (Budget)", "Aantal verschil",
    "PDF Prijs p.e.", "NETTO Prijs p.e.", "Prijs verschil €", "Prijs verschil %",
    "NETTO Art.nr. Jumbo", "NETTO Manual nr.", "NETTO Omschrijving",
    "IB Nummer", "IB Manual", "IB Omschrijving", "IB Prijs (Budget)",
    "IB Match methode", "Pagina (PDF)",
]
_COL_WIDTHS = [12, 48, 20, 15, 15, 10, 10, 14, 13, 14, 14, 14, 13, 16, 16, 38, 16, 14, 42, 14, 14, 8]


def _write_budget_sheet(wb, budget_groups: list):
    """Add a 'Budget' sheet with grouped PDF totals compared against Budget reference rows."""
    from openpyxl.styles import Border, Side
    ws = wb.create_sheet("Budget")

    F_GROUP   = PatternFill("solid", fgColor="2E75B6")
    F_PDF_HDR = PatternFill("solid", fgColor="4472C4")
    F_ITEM_A  = PatternFill("solid", fgColor="FFFFFF")
    F_ITEM_B  = PatternFill("solid", fgColor="F2F2F2")
    F_TOTAL   = PatternFill("solid", fgColor="DCE6F1")
    F_BUD_HDR = PatternFill("solid", fgColor="375623")
    F_BUD_VAL = PatternFill("solid", fgColor="E2EFDA")
    F_DIFF_OK = PatternFill("solid", fgColor="C6EFCE")
    F_DIFF_HI = PatternFill("solid", fgColor="FFC7CE")
    F_DIFF_LO = PatternFill("solid", fgColor="FFEB9C")
    F_DIFF_HDR= PatternFill("solid", fgColor="595959")

    WHITE  = Font(bold=True, color="FFFFFF", size=11)
    BOLD11 = Font(bold=True, size=11)
    BOLD14 = Font(bold=True, size=14)
    THICK  = Border(bottom=Side(style="medium"))
    THIN   = Border(bottom=Side(style="thin"))
    CTR    = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    col_widths = [14, 65, 16, 8, 10, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def cell(r, c, val, fill=None, font=None, align=None, border=None, fmt=None):
        ce = ws.cell(row=r, column=c, value=val)
        if fill:   ce.fill      = fill
        if font:   ce.font      = font
        if align:  ce.alignment = align
        if border: ce.border    = border
        if fmt:    ce.number_format = fmt
        return ce

    def hdr_row(r, values, fill, font):
        for c, v in enumerate(values, 1):
            cell(r, c, v, fill=fill, font=font, align=CTR, border=THICK)

    # Page title
    cell(1, 1, "Budget Vergelijking – Gegroepeerde Offerte Items", font=BOLD14)
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 24

    current_row = 3

    for grp in budget_groups:
        br = grp["budget_row"]

        # Group banner
        cell(current_row, 1, grp["name"], fill=F_GROUP, font=Font(bold=True, color="FFFFFF", size=13))
        ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=7)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        cell(current_row, 1, grp["label"], fill=F_TOTAL, font=Font(italic=True, size=9))
        ws.merge_cells(start_row=current_row, end_row=current_row, start_column=1, end_column=7)
        current_row += 1

        # PDF items header
        hdr_row(current_row,
                ["Art.nr. (PDF)", "Omschrijving (PDF)", "Manual Ref", "Pagina",
                 "Aantal", "Prijs p.e.", "Totaal"],
                fill=F_PDF_HDR, font=WHITE)
        current_row += 1

        pdf_total_qty = 0.0
        for idx, pi in enumerate(grp["pdf_items"]):
            f = F_ITEM_A if idx % 2 == 0 else F_ITEM_B
            cell(current_row, 1, pi.art_nr,      fill=f, font=None, align=LEFT)
            cell(current_row, 2, pi.description, fill=f, font=None, align=LEFT)
            cell(current_row, 3, pi.manual_ref,  fill=f, font=None, align=LEFT)
            cell(current_row, 4, pi.page,        fill=f, font=None, align=CTR)
            cell(current_row, 5, pi.quantity,    fill=f, font=None, align=CTR, fmt="#,##0.##")
            cell(current_row, 6, pi.unit_price,  fill=f, font=None, align=CTR, fmt="#,##0.00")
            cell(current_row, 7, pi.total,       fill=f, font=None, align=CTR, fmt="#,##0.00")
            pdf_total_qty += pi.quantity
            current_row += 1

        # PDF subtotal
        for c in range(1, 8):
            ws.cell(current_row, c).fill   = F_TOTAL
            ws.cell(current_row, c).border = THIN
        cell(current_row, 1, "TOTAAL PDF", fill=F_TOTAL, font=BOLD11, align=LEFT)
        cell(current_row, 5, pdf_total_qty,     fill=F_TOTAL, font=BOLD11, align=CTR, fmt="#,##0.##")
        cell(current_row, 7, grp["pdf_total"],  fill=F_TOTAL, font=BOLD11, align=CTR, fmt="#,##0.00")
        current_row += 2

        # Budget reference
        hdr_row(current_row,
                ["Budget Rij", "Artikelnaam (Budget)", "Manual", "Eenheid",
                 "Aantal (Budget)", "Prijs p.e. (Budget)", "Totaal (Budget)"],
                fill=F_BUD_HDR, font=WHITE)
        current_row += 1

        cell(current_row, 1, f"Rij {br['row_num']}",  fill=F_BUD_VAL, align=CTR)
        cell(current_row, 2, br["artikelnaam"],         fill=F_BUD_VAL, align=LEFT)
        cell(current_row, 3, br["manual"],              fill=F_BUD_VAL, align=CTR)
        cell(current_row, 4, br["eenheid"],             fill=F_BUD_VAL, align=CTR)
        cell(current_row, 5, br["aantal"],              fill=F_BUD_VAL, align=CTR, fmt="#,##0.##")
        cell(current_row, 6, br["prijs"],               fill=F_BUD_VAL, align=CTR, fmt="#,##0.00")
        cell(current_row, 7, br["totaal"],              fill=F_BUD_VAL, align=CTR, fmt="#,##0.00")
        current_row += 2

        # Difference summary
        hdr_row(current_row,
                ["VERGELIJKING", "PDF Totaal", "Budget Totaal",
                 "Verschil (EUR)", "Verschil (%)", "", ""],
                fill=F_DIFF_HDR, font=WHITE)
        current_row += 1

        d = grp["diff_eur"]
        diff_fill = F_DIFF_OK if abs(d) < 100 else (F_DIFF_HI if d > 0 else F_DIFF_LO)
        for c in range(1, 8):
            ws.cell(current_row, c).fill = diff_fill
        cell(current_row, 1, grp["name"],          fill=diff_fill, font=BOLD11, align=CTR)
        cell(current_row, 2, grp["pdf_total"],      fill=diff_fill, font=BOLD11, align=CTR, fmt="#,##0.00")
        cell(current_row, 3, br["totaal"],          fill=diff_fill, font=BOLD11, align=CTR, fmt="#,##0.00")
        cell(current_row, 4, grp["diff_eur"],       fill=diff_fill, font=BOLD11, align=CTR, fmt="+#,##0.00;-#,##0.00")
        pct = grp["diff_pct"]
        cell(current_row, 5, f"{pct:+.1f}%" if pct is not None else "",
             fill=diff_fill, font=BOLD11, align=CTR)
        ws.row_dimensions[current_row].height = 20
        current_row += 4


def export_excel(results: List[MatchResult], output_path: str, budget_groups: list = None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vergelijking"

    # Header row
    for col, (h, w) in enumerate(zip(_HEADERS, _COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    for row_num, mr in enumerate(results, 2):
        pi = mr.pdf_item
        ni = mr.netto_item
        ii_list = mr.ib_items
        ii = ii_list[0] if ii_list else None

        ib_qty    = mr.ib_qty

        pct_diff = ""
        if mr.price_diff is not None and ni and ni.netto_price:
            pct_diff = f"{mr.price_diff / ni.netto_price * 100:+.1f}%"

        row_data = [
            pi.art_nr         if pi else "",
            pi.description    if pi else "",
            pi.section        if pi else "",
            pi.manual_ref     if pi else "",
            mr.pdf_match_method,
            f"{mr.confidence:.0%}" if mr.confidence else "",
            pi.quantity       if pi else "",
            ib_qty            if ib_qty is not None else "",
            mr.qty_diff       if mr.qty_diff is not None else "",
            pi.unit_price     if pi else "",
            ni.netto_price    if ni else "",
            mr.price_diff     if mr.price_diff is not None else "",
            pct_diff,
            ni.art_nr_jumbo   if ni else "",
            ni.manual_nr      if ni else "",
            ni.description    if ni else "",
            ii.nummer         if ii else "",
            ii.manual         if ii else "",
            ii.description    if ii else "",
            ii.price          if ii else "",
            mr.ib_match_method,
            pi.page           if pi else "",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=val)

        # Row-level background
        if mr.pdf_match_method == "unmatched" and not ii_list:
            fill = _FILL_YELLOW   # PDF item with no match anywhere
        else:
            fill = None

        if fill:
            for col in range(1, len(_HEADERS) + 1):
                ws.cell(row=row_num, column=col).fill = fill

        # Cell-level highlights (override row fill)
        if mr.qty_diff is not None and mr.qty_diff != 0:
            ws.cell(row=row_num, column=9).fill  = _FILL_ORANGE   # Aantal verschil
        if mr.price_diff is not None and abs(mr.price_diff) > 0.01:
            ws.cell(row=row_num, column=12).fill = _FILL_RED       # Prijs verschil €
            ws.cell(row=row_num, column=13).fill = _FILL_RED       # Prijs verschil %

    # Summary sheet
    ws2 = wb.create_sheet("Samenvatting")
    pdf_total       = sum(1 for r in results if r.pdf_item)
    matched_exact   = sum(1 for r in results if r.pdf_match_method == "exact_artnr")
    matched_manual  = sum(1 for r in results if r.pdf_match_method == "manual_code")
    matched_fuzzy   = sum(1 for r in results if r.pdf_match_method == "fuzzy")
    unmatched_pdf = sum(1 for r in results if r.pdf_match_method == "unmatched" and r.pdf_item)
    ib_exact      = sum(1 for r in results if r.ib_match_method == "exact_artnr")
    ib_fuzzy      = sum(1 for r in results if r.ib_match_method == "fuzzy")
    ib_none       = sum(1 for r in results if r.pdf_item and not r.ib_items)
    qty_diffs     = sum(1 for r in results if r.qty_diff   is not None and r.qty_diff   != 0)
    price_diffs   = sum(1 for r in results if r.price_diff is not None and abs(r.price_diff) > 0.01)

    rows = [
        ("Resultaten vergelijking", ""),
        ("", ""),
        ("PDF items totaal",                        pdf_total),
        ("  Gematch NETTO via art.nr. (exact)",     matched_exact),
        ("  Gematch NETTO via Manual code",          matched_manual),
        ("  Gematch NETTO via fuzzy omschrijving",  matched_fuzzy),
        ("  Niet gematch NETTO",                     unmatched_pdf),
        ("", ""),
        ("IB Budget koppeling",                      ""),
        ("  Exact art.nr. (col BA)",                 ib_exact),
        ("  Fuzzy omschrijving",                     ib_fuzzy),
        ("  Geen IB match",                          ib_none),
        ("", ""),
        ("Afwijkingen",                              ""),
        ("  Aantal verschil (PDF ≠ IB Budget)",      qty_diffs),
        ("  Prijs verschil  (PDF ≠ NETTO)",          price_diffs),
    ]
    for r, (lbl, val) in enumerate(rows, 1):
        c = ws2.cell(row=r, column=1, value=lbl)
        c.font = _FONT_BOLD
        ws2.cell(row=r, column=2, value=val)
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 12

    if budget_groups:
        _write_budget_sheet(wb, budget_groups)

    wb.save(output_path)


# ─── Tkinter Application ───────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Offerte Vergelijker – Van Keulen")
        self.geometry("1280x800")
        self.minsize(900, 600)
        self._results: List[MatchResult] = []
        self._budget_groups: list = []
        self._logo_photo = None   # keep reference so GC doesn't collect it
        self._apply_logo()
        self._build_ui()

    # ── Logo ───────────────────────────────────────────────────────────────

    def _apply_logo(self):
        if not HAS_PIL:
            return
        try:
            from PIL import ImageTk
            icon_img = _make_jumbo_logo(64, 64)
            self._icon_photo = ImageTk.PhotoImage(icon_img)
            self.iconphoto(True, self._icon_photo)
            # larger version for header bar
            header_img = _make_jumbo_logo(160, 52)
            self._logo_photo = ImageTk.PhotoImage(header_img)
        except Exception:
            pass

    # ── UI construction ────────────────────────────────────────────────────

    def _build_ui(self):
        # Header bar with logo
        if self._logo_photo:
            frm_header = tk.Frame(self, bg="#FDC400", height=58)
            frm_header.pack(fill="x")
            frm_header.pack_propagate(False)
            tk.Label(frm_header, image=self._logo_photo, bg="#FDC400").pack(side="right", padx=10, pady=4)
            tk.Label(
                frm_header,
                text="Offerte Vergelijker",
                bg="#FDC400", fg="#1A1A1A",
                font=("Arial", 16, "bold"),
            ).pack(side="left", padx=14, pady=4)

        # File selection panel
        frm_files = ttk.LabelFrame(self, text="Bestanden", padding=8)
        frm_files.pack(fill="x", padx=10, pady=6)

        self._v_pdf    = tk.StringVar()
        self._v_netto  = tk.StringVar()
        self._v_budget = tk.StringVar()
        self._v_output = tk.StringVar()

        self._auto_fill_paths()

        file_rows = [
            ("PDF Offerte:",      self._v_pdf,    [("PDF bestanden", "*.pdf")]),
            ("NETTO Prijslijst:", self._v_netto,  [("Excel bestanden", "*.xlsx")]),
            ("Budget (IB):",      self._v_budget, [("Excel macro bestanden", "*.xlsm"), ("Excel bestanden", "*.xlsx")]),
            ("Output Excel:",     self._v_output, [("Excel bestanden", "*.xlsx")]),
        ]
        for i, (label, var, ftypes) in enumerate(file_rows):
            ttk.Label(frm_files, text=label, width=18, anchor="e").grid(
                row=i, column=0, sticky="e", pady=2, padx=(0, 4)
            )
            ttk.Entry(frm_files, textvariable=var, width=90).grid(
                row=i, column=1, sticky="ew", padx=4
            )
            ttk.Button(
                frm_files, text="…",
                command=lambda v=var, f=ftypes: self._browse(v, f)
            ).grid(row=i, column=2)
        frm_files.columnconfigure(1, weight=1)

        # Action bar
        frm_btns = ttk.Frame(self)
        frm_btns.pack(fill="x", padx=10, pady=2)
        self._btn_run = ttk.Button(frm_btns, text="▶  Analyseren", command=self._run)
        self._btn_run.pack(side="left", padx=4)
        self._btn_export = ttk.Button(
            frm_btns, text="💾  Exporteer Excel", command=self._export, state="disabled"
        )
        self._btn_export.pack(side="left", padx=4)
        self._lbl_status = ttk.Label(frm_btns, text="", foreground="gray")
        self._lbl_status.pack(side="left", padx=10)

        # Legend
        frm_legend = ttk.Frame(self)
        frm_legend.pack(fill="x", padx=10)
        legend = [
            ("#C6EFCE", "Overeenkomst"),
            ("#FFCC00", "Aantal verschil"),
            ("#FFEB9C", "Niet gematch (PDF)"),
            ("#FFFFFF", "▲ = Prijs verschil (zie kolom Prijs Δ)"),
        ]
        for color, label in legend:
            tk.Label(frm_legend, text="  ", bg=color, relief="solid", bd=1).pack(side="left", padx=(8, 2))
            ttk.Label(frm_legend, text=label).pack(side="left", padx=(0, 8))

        # Notebook
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=10, pady=6)
        self._nb = nb

        self._build_tab_results(nb)
        self._build_tab_review(nb)
        self._build_tab_log(nb)

    def _build_tab_results(self, nb):
        frm = ttk.Frame(nb)
        nb.add(frm, text="Resultaten")

        cols = (
            "art_nr", "manual_nr", "omschrijving", "sectie",
            "pdf_qty", "ib_qty", "qty_diff",
            "pdf_prijs", "netto_prijs", "prijs_diff",
            "methode", "confidence",
        )
        headings = (
            "Art.nr.", "Manual nr.", "Omschrijving (PDF)", "Sectie",
            "PDF Aant.", "IB Aant.", "Verschil",
            "PDF Prijs", "NETTO Prijs", "Prijs Δ",
            "Methode", "Match %",
        )
        widths = (90, 100, 340, 130, 75, 80, 70, 90, 90, 90, 110, 70)

        sb_y = ttk.Scrollbar(frm, orient="vertical")
        sb_x = ttk.Scrollbar(frm, orient="horizontal")
        self._tv = ttk.Treeview(
            frm, columns=cols, show="headings",
            yscrollcommand=sb_y.set, xscrollcommand=sb_x.set,
        )
        sb_y.config(command=self._tv.yview)
        sb_x.config(command=self._tv.xview)

        for col, heading, w in zip(cols, headings, widths):
            self._tv.heading(col, text=heading, command=lambda c=col: self._sort(c))
            self._tv.column(col, width=w, minwidth=40)

        self._tv.tag_configure("ok",         background="#C6EFCE")
        self._tv.tag_configure("qty_diff",   background="#FFCC00")
        self._tv.tag_configure("price_diff")                        # no row color — column value shows ▲
        self._tv.tag_configure("both_diff",  background="#FFCC00")  # qty color only, price shown via ▲
        self._tv.tag_configure("unmatched",  background="#FFEB9C")

        self._tv.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)

    def _build_tab_review(self, nb):
        frm = ttk.Frame(nb)
        nb.add(frm, text="Te controleren")

        ttk.Label(
            frm,
            text="Items die handmatig gecontroleerd moeten worden (fuzzy match < 90% of niet gematch):",
        ).pack(anchor="w", padx=6, pady=3)

        cols = ("art_nr", "pdf_desc", "pdf_qty", "netto_desc", "netto_prijs", "confidence", "methode")
        headings = ("Art.nr.", "PDF Omschrijving", "PDF Qty", "NETTO Omschrijving", "NETTO Prijs", "Match %", "Methode")
        widths   = (90, 310, 75, 310, 100, 70, 100)

        inner = ttk.Frame(frm)
        inner.pack(fill="both", expand=True)

        sb = ttk.Scrollbar(inner, orient="vertical")
        self._tv_rev = ttk.Treeview(
            inner, columns=cols, show="headings", yscrollcommand=sb.set
        )
        sb.config(command=self._tv_rev.yview)

        for col, heading, w in zip(cols, headings, widths):
            self._tv_rev.heading(col, text=heading)
            self._tv_rev.column(col, width=w, minwidth=40)

        self._tv_rev.tag_configure("fuzzy",     background="#FFEB9C")
        self._tv_rev.tag_configure("unmatched", background="#FFC7CE")

        self._tv_rev.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        inner.rowconfigure(0, weight=1)
        inner.columnconfigure(0, weight=1)

    def _build_tab_log(self, nb):
        frm = ttk.Frame(nb)
        nb.add(frm, text="Log")
        self._log = scrolledtext.ScrolledText(
            frm, state="disabled", font=("Consolas", 9), height=8
        )
        self._log.pack(fill="both", expand=True)

    # ── Helpers ────────────────────────────────────────────────────────────

    def _auto_fill_paths(self):
        folder = DEFAULT_FOLDER
        if not folder.exists():
            return
        pdf_files   = sorted(folder.glob("*.pdf"))
        netto_files = [f for f in folder.glob("*.xlsx")
                       if any(k in f.name for k in ("NETTO", "netto", "Prijslijst", "prijslijst"))]
        budget_files = [f for f in folder.glob("*.xlsm")]
        if pdf_files:    self._v_pdf.set(str(pdf_files[0]))
        if netto_files:  self._v_netto.set(str(netto_files[0]))
        if budget_files: self._v_budget.set(str(budget_files[0]))
        self._v_output.set(str(folder / "Vergelijking_output.xlsx"))

    def _browse(self, var: tk.StringVar, ftypes):
        path = filedialog.askopenfilename(filetypes=ftypes + [("Alle bestanden", "*.*")])
        if path:
            var.set(path)

    def _log_msg(self, msg: str):
        self._log.config(state="normal")
        self._log.insert("end", msg + "\n")
        self._log.see("end")
        self._log.config(state="disabled")
        self.update_idletasks()

    def _sort(self, col: str):
        """Sort treeview by column (toggle asc/desc)."""
        items = [(self._tv.set(k, col), k) for k in self._tv.get_children("")]
        try:
            items.sort(key=lambda t: float(t[0]) if t[0] else 0, reverse=False)
        except ValueError:
            items.sort(key=lambda t: t[0].lower())
        for idx, (_, k) in enumerate(items):
            self._tv.move(k, "", idx)

    # ── Main actions ───────────────────────────────────────────────────────

    def _run(self):
        pdf_path   = self._v_pdf.get()
        netto_path = self._v_netto.get()

        if not all([pdf_path, netto_path]):
            messagebox.showerror("Fout", "Selecteer minimaal PDF en NETTO bestanden.")
            return

        self._btn_run.config(state="disabled")
        self._lbl_status.config(text="Bezig met analyseren…")

        try:
            self._log_msg(f"PDF lezen:   {pdf_path}")
            pdf_items = parse_pdf(pdf_path)
            self._log_msg(f"  → {len(pdf_items)} items gevonden in PDF")

            self._log_msg(f"NETTO lezen: {netto_path}")
            netto_items = read_netto(netto_path)
            self._log_msg(f"  → {len(netto_items)} items in NETTO lijst")

            if not HAS_RAPIDFUZZ:
                self._log_msg("⚠  rapidfuzz niet geïnstalleerd – fuzzy matching uitgeschakeld.")
                self._log_msg("   Installeer met: pip install rapidfuzz")

            budget_path = self._v_budget.get()
            ib_items = []
            if budget_path:
                self._log_msg(f"IB Budget lezen: {budget_path}")
                ib_items = read_ib_items(budget_path)
                self._log_msg(f"  → {len(ib_items)} IB Budget items geladen (rijen 1440-2763)")
                self._budget_groups = build_budget_groups(pdf_items, budget_path)
                for g in self._budget_groups:
                    self._log_msg(
                        f"  → Groep '{g['name']}': {len(g['pdf_items'])} items  "
                        f"PDF totaal € {g['pdf_total']:,.2f}  "
                        f"Budget totaal € {g['budget_row']['totaal']:,.2f}  "
                        f"Verschil € {g['diff_eur']:+,.2f}"
                    )
            else:
                self._budget_groups = []

            self._log_msg("Matchen…")
            self._results = build_matches(pdf_items, netto_items, ib_items)

            n_exact   = sum(1 for r in self._results if r.pdf_match_method == "exact_artnr")
            n_manual  = sum(1 for r in self._results if r.pdf_match_method == "manual_code")
            n_fuzzy   = sum(1 for r in self._results if r.pdf_match_method == "fuzzy")
            n_unmatch = sum(1 for r in self._results if r.pdf_match_method == "unmatched" and r.pdf_item)
            n_qty     = sum(1 for r in self._results if r.qty_diff   is not None and r.qty_diff   != 0)
            n_price   = sum(1 for r in self._results if r.price_diff is not None and abs(r.price_diff) > 0.01)

            self._log_msg(
                f"  → exact: {n_exact}  |  manual: {n_manual}  |  "
                f"fuzzy: {n_fuzzy}  |  ongematch: {n_unmatch}"
            )
            self._log_msg(f"  → {n_qty} aantal-afwijkingen  |  {n_price} prijs-afwijkingen")

            self._populate_results()
            self._populate_review()
            self._btn_export.config(state="normal")
            self._lbl_status.config(
                text=f"✓ {len(self._results)} rijen  |  "
                     f"{n_qty} aant.verschil  |  {n_price} prijsverschil"
            )
            self._nb.select(0)   # show Results tab

        except Exception as exc:
            messagebox.showerror("Fout", str(exc))
            import traceback
            traceback.print_exc()
        finally:
            self._btn_run.config(state="normal")

    def _populate_results(self):
        for k in self._tv.get_children():
            self._tv.delete(k)

        # Group results by art_nr, preserving order of first occurrence
        seen: dict = {}
        ordered_keys = []
        for mr in self._results:
            if not mr.pdf_item:
                continue
            key = mr.pdf_item.art_nr
            if key not in seen:
                seen[key] = []
                ordered_keys.append(key)
            seen[key].append(mr)

        for key in ordered_keys:
            mrs = seen[key]
            first = mrs[0]
            pi = first.pdf_item
            ni = first.netto_item

            # Aggregate quantities and combine unique sections
            total_pdf_qty = sum(mr.pdf_item.quantity for mr in mrs)
            unique_sections = []
            for mr in mrs:
                s = mr.pdf_item.section if mr.pdf_item else ""
                if s and s not in unique_sections:
                    unique_sections.append(s)
            sectie = " / ".join(unique_sections)

            ib_qty_val = first.ib_qty
            qty_diff_val = round(total_pdf_qty - ib_qty_val, 4) if ib_qty_val is not None else None

            art_nr    = pi.art_nr
            manual_nr = ni.manual_nr      if ni else ""
            desc      = (pi.description[:55] + "…") if len(pi.description) > 55 else pi.description
            pdf_qty   = total_pdf_qty
            ib_qty    = ib_qty_val        if ib_qty_val  is not None else ""
            qty_diff  = qty_diff_val      if qty_diff_val is not None else ""
            pdf_p     = f"€ {pi.unit_price:,.2f}"  if pi else ""
            netto_p   = f"€ {ni.netto_price:,.2f}" if ni else ""
            price_diff = first.price_diff
            prijs_d   = (
                f"▲ € {price_diff:+,.2f}"
                if price_diff is not None else ""
            )
            methode   = first.pdf_match_method
            conf      = f"{first.confidence:.0%}" if first.confidence else ""

            has_qty   = qty_diff_val  is not None and qty_diff_val  != 0
            has_price = price_diff    is not None and abs(price_diff) > 0.01

            if first.pdf_match_method == "unmatched":
                tag = "unmatched"
            elif has_qty and has_price:
                tag = "both_diff"
            elif has_qty:
                tag = "qty_diff"
            elif has_price:
                tag = "price_diff"
            else:
                tag = "ok"

            self._tv.insert(
                "", "end", tags=(tag,),
                values=(art_nr, manual_nr, desc, sectie, pdf_qty, ib_qty, qty_diff,
                        pdf_p, netto_p, prijs_d, methode, conf),
            )

    def _populate_review(self):
        for k in self._tv_rev.get_children():
            self._tv_rev.delete(k)

        for mr in self._results:
            needs_review = (
                (mr.pdf_match_method == "fuzzy" and mr.confidence < 0.90)
                or (mr.pdf_match_method == "unmatched" and mr.pdf_item)
            )
            if not needs_review:
                continue
            pi = mr.pdf_item
            ni = mr.netto_item
            tag = "fuzzy" if mr.pdf_match_method == "fuzzy" else "unmatched"
            pdf_d   = (pi.description[:55] + "…") if len(pi.description) > 55 else pi.description
            netto_d = (
                (ni.description[:55] + "…") if ni and len(ni.description) > 55
                else (ni.description if ni else "---")
            )
            self._tv_rev.insert("", "end", tags=(tag,), values=(
                pi.art_nr,
                pdf_d,
                pi.quantity,
                netto_d,
                f"€ {ni.netto_price:,.2f}" if ni else "---",
                f"{mr.confidence:.0%}" if mr.confidence else "0%",
                mr.pdf_match_method,
            ))

    def _export(self):
        out = self._v_output.get()
        if not out:
            out = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel bestanden", "*.xlsx")],
            )
            if not out:
                return
            self._v_output.set(out)
        try:
            export_excel(self._results, out, budget_groups=self._budget_groups or None)
            self._log_msg(f"Export opgeslagen: {out}")
            messagebox.showinfo("Klaar", f"Excel opgeslagen:\n{out}")
        except Exception as exc:
            messagebox.showerror("Fout bij exporteren", str(exc))


# ─── Entry point ───────────────────────────────────────────────────────────────

def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
