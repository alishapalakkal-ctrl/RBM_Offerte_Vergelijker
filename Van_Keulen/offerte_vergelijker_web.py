#!/usr/bin/env python3
"""
Offerte Vergelijker – Streamlit web app
Run locally:  streamlit run sloopwerk/offerte_vergelijker_web.py
Host on:      https://streamlit.io/cloud  (free, connect GitHub repo)
"""

import io
import os
import re
import tempfile
from dataclasses import dataclass, field
from typing import Dict, List, Optional

import pandas as pd
import streamlit as st

# ── Optional deps ──────────────────────────────────────────────────────────────
try:
    import pdfplumber
except ImportError:
    st.error("pdfplumber niet gevonden — `pip install pdfplumber`")
    st.stop()

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl niet gevonden — `pip install openpyxl`")
    st.stop()

try:
    from rapidfuzz import fuzz, process as rfprocess
    HAS_RAPIDFUZZ = True
except ImportError:
    HAS_RAPIDFUZZ = False

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Offerte Vergelijker – Van Keulen",
    page_icon="🟡",
    layout="wide",
)

st.markdown("""
<style>
  .jumbo-hdr {
      background: linear-gradient(135deg, #FDC400 0%, #e8ac00 100%);
      padding: 16px 28px; border-radius: 12px; margin-bottom: 20px;
      display: flex; align-items: center; gap: 18px;
      box-shadow: 0 3px 10px rgba(0,0,0,.15);
  }
  .jumbo-hdr h1 { margin: 0; font-size: 28px; font-weight: 800; color: #1a1a1a; }
  .jumbo-hdr p  { margin: 4px 0 0; font-size: 13px; color: #444; }
  [data-testid="stSidebar"] { background: #f5f5f5; }
  [data-testid="stMetricValue"] { font-size: 28px !important; }
  .legend-row { display:flex; gap:18px; margin: 6px 0 14px; font-size:13px; }
  .legend-chip { padding: 3px 12px; border-radius: 5px; font-weight:600; }
</style>
""", unsafe_allow_html=True)


# ─── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class PdfItem:
    art_nr: str
    description: str
    quantity: float
    unit_price: float
    total: float
    section: str = ""
    manual_ref: str = ""
    page: int = 0


@dataclass
class NettoItem:
    art_nr_jumbo: str
    manual_nr: str
    art_nr_leverancier: str
    groep: str
    description: str
    netto_price: float
    unit: str = ""


@dataclass
class IBItem:
    nummer: str
    description: str
    manual: str
    unit: str
    quantity: float
    price: float
    pdf_art_nr: str


@dataclass
class MatchResult:
    pdf_item: Optional[PdfItem] = None
    netto_item: Optional[NettoItem] = None
    ib_items: List[IBItem] = field(default_factory=list)
    pdf_match_method: str = "unmatched"
    ib_match_method: str = "unmatched"
    confidence: float = 0.0

    @property
    def ib_qty(self) -> Optional[float]:
        return sum(i.quantity for i in self.ib_items) if self.ib_items else None

    @property
    def price_diff(self) -> Optional[float]:
        if self.pdf_item and self.netto_item and self.netto_item.netto_price:
            return round(self.pdf_item.unit_price - self.netto_item.netto_price, 4)
        return None


# ─── PDF Parser ────────────────────────────────────────────────────────────────

_ITEM_END = re.compile(
    r'\s+((?:\d{1,3}\.)*\d+)\s+((?:\d{1,3}\.)*\d+,\d{2})\s+((?:\d{1,3}\.)*\d+,\d{2})\s*$'
)
_ART_NR  = re.compile(r'^(\d{3,7})\s+(.*)')
_MANUAL  = re.compile(r'^Manual\s+(\S+)', re.IGNORECASE)
_POSTAL  = re.compile(r'^\d{4}\s+[A-Z]{2}\s+')
_SECTION = re.compile(r'^[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ /\-]{1,35}$')
_SKIP_SW = (
    'Offertenummer','Projectnummer','Ontvangende','Art.nr.','Art. nr.',
    'Op alle','Van Keulen','Postbus','www.','Pagina ',
    'Leverdatum','Uw contact','Planning','Aanvangsdatum',
    'EUR ','FSC ','PEFC ','ISO','Met vriendelijke',
    'Naam in blokletters','Handtekening','Totaalbedrag','Dit document',
    'Opdrachtgever','Jumbo Supermarkten','Mevr.','Dhr.','Nederland',
    'Offertedatum','Vervaldatum','Verkoper','Uw referentie',
)


def _dutch(s: str) -> float:
    return float(s.replace('.', '').replace(',', '.'))


def _skip(line: str) -> bool:
    if not line or len(line) <= 1:
        return True
    if _POSTAL.match(line):
        return True
    return any(line.startswith(s) for s in _SKIP_SW)


def parse_pdf(path) -> List[PdfItem]:
    items: List[PdfItem] = []
    section = manual_ref = art_nr = ""
    lines: List[str] = []
    page_nr = 0

    def flush():
        nonlocal art_nr, lines
        if not art_nr:
            return
        text = " ".join(lines).strip()
        m = _ITEM_END.search(text)
        if m:
            desc = text[:m.start()].strip()
            qty, price, total = _dutch(m.group(1)), _dutch(m.group(2)), _dutch(m.group(3))
            if price > 0:
                items.append(PdfItem(
                    art_nr=art_nr, description=desc, quantity=qty,
                    unit_price=price, total=total,
                    section=section, manual_ref=manual_ref, page=page_nr,
                ))
        art_nr = ""
        lines.clear()

    with pdfplumber.open(path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            if page_num == 1:
                continue
            page_nr = page_num
            text = page.extract_text()
            if not text:
                continue
            for raw in text.split('\n'):
                line = raw.strip()
                if _skip(line):
                    continue
                m = _MANUAL.match(line)
                if m:
                    flush()
                    manual_ref = m.group(1)
                    continue
                if line.startswith('['):
                    continue
                m = _ART_NR.match(line)
                if m:
                    if art_nr and not _ITEM_END.search(" ".join(lines)):
                        lines.append(line)
                        if _ITEM_END.search(" ".join(lines)):
                            flush()
                    else:
                        flush()
                        art_nr = m.group(1)
                        lines = [m.group(2)]
                    continue
                if art_nr:
                    lines.append(line)
                    if _ITEM_END.search(" ".join(lines)):
                        flush()
                elif line and _SECTION.match(line):
                    section = line
        flush()
    return items


# ─── Excel Readers ─────────────────────────────────────────────────────────────

def read_netto(src) -> List[NettoItem]:
    df = pd.read_excel(src, sheet_name=0, header=None)
    items: List[NettoItem] = []
    for _, row in df.iloc[2:].iterrows():
        def cell(i):
            return str(row.iloc[i]).strip() if pd.notna(row.iloc[i]) else ""
        jumbo = cell(0)
        if not jumbo or jumbo in ('nan', 'None'):
            continue
        art_lev = cell(2)
        if art_lev in ('nan', 'None', 'Geen artikelno.', 'Niet leverbaar'):
            art_lev = ""
        price = 0.0
        for ci in (16, 5):
            try:
                price = float(row.iloc[ci])
                break
            except (ValueError, TypeError):
                pass
        items.append(NettoItem(
            art_nr_jumbo=jumbo, manual_nr=cell(1), art_nr_leverancier=art_lev,
            groep=cell(3), description=cell(4), netto_price=price, unit=cell(11),
        ))
    return items


def read_ib_items(src, row_start: int = 1440, row_end: int = 2763) -> List[IBItem]:
    df = pd.read_excel(src, sheet_name="Budget", header=None, engine="openpyxl")
    items: List[IBItem] = []
    for idx in range(row_start - 1, min(row_end, len(df))):
        row = df.iloc[idx]
        def cell(i):
            return str(row.iloc[i]).strip() if i < len(row) and pd.notna(row.iloc[i]) else ""
        if cell(5) != "2":
            continue
        desc = cell(7)
        if not desc or desc in ("nan", "NaT", "0"):
            continue
        try:    qty   = float(row.iloc[10])
        except: qty   = 0.0
        try:    price = float(row.iloc[11])
        except: price = 0.0
        pdf_nr = cell(52) if len(row) > 52 else ""
        if pdf_nr in ("nan", "NaT"):
            pdf_nr = ""
        items.append(IBItem(
            nummer=cell(6), description=desc, manual=cell(8), unit=cell(9),
            quantity=qty, price=price, pdf_art_nr=pdf_nr,
        ))
    return items


# ─── Matcher ───────────────────────────────────────────────────────────────────

def build_matches(
    pdf_items: List[PdfItem],
    netto_items: List[NettoItem],
    ib_items: List[IBItem],
    fuzzy_threshold: int = 70,
) -> List[MatchResult]:
    by_artnr:  Dict[str, NettoItem]        = {n.art_nr_leverancier: n for n in netto_items if n.art_nr_leverancier}
    by_manual: Dict[str, List[NettoItem]]  = {}
    for n in netto_items:
        k = n.manual_nr.strip().upper()
        if k:
            by_manual.setdefault(k, []).append(n)

    ib_by_pdf = {ii.pdf_art_nr: ii for ii in ib_items if ii.pdf_art_nr}
    ib_descs  = [ii.description for ii in ib_items]  if HAS_RAPIDFUZZ else []
    n_descs   = [n.description  for n  in netto_items] if HAS_RAPIDFUZZ else []

    results: List[MatchResult] = []
    for pi in pdf_items:
        mr = MatchResult(pdf_item=pi)

        ni = by_artnr.get(pi.art_nr)
        if ni:
            mr.netto_item = ni
            mr.pdf_match_method = "exact_artnr"
            mr.confidence = 1.0

        if not mr.netto_item and pi.manual_ref:
            cands = by_manual.get(pi.manual_ref.strip().upper(), [])
            if cands:
                if HAS_RAPIDFUZZ and len(cands) > 1:
                    ds   = [c.description for c in cands]
                    best = rfprocess.extractOne(pi.description, ds, scorer=fuzz.token_set_ratio)
                    mr.netto_item = cands[ds.index(best[0])] if best else cands[0]
                else:
                    mr.netto_item = cands[0]
                mr.pdf_match_method = "manual_code"
                mr.confidence = 0.85

        if not mr.netto_item and HAS_RAPIDFUZZ:
            best = rfprocess.extractOne(
                pi.description, n_descs,
                scorer=fuzz.token_set_ratio, score_cutoff=fuzzy_threshold,
            )
            if best:
                mr.netto_item = netto_items[n_descs.index(best[0])]
                mr.pdf_match_method = "fuzzy"
                mr.confidence = best[1] / 100.0

        ii = ib_by_pdf.get(pi.art_nr)
        if ii:
            mr.ib_items = [ii]
            mr.ib_match_method = "exact_artnr"
        elif HAS_RAPIDFUZZ and ib_descs:
            best = rfprocess.extractOne(
                pi.description, ib_descs,
                scorer=fuzz.token_set_ratio, score_cutoff=fuzzy_threshold,
            )
            if best:
                mr.ib_items = [ib_items[ib_descs.index(best[0])]]
                mr.ib_match_method = "fuzzy"

        results.append(mr)
    return results


# ─── Results → DataFrame ───────────────────────────────────────────────────────

_STATUS_COLOR = {
    "ok":         "#C6EFCE",
    "qty_diff":   "#FFCC00",
    "price_diff": "#FFF2CC",
    "unmatched":  "#FFC7CE",
}


def results_to_df(results: List[MatchResult]) -> pd.DataFrame:
    seen:  Dict[str, List[MatchResult]] = {}
    order: List[str] = []
    for mr in results:
        if not mr.pdf_item:
            continue
        k = mr.pdf_item.art_nr
        if k not in seen:
            seen[k] = []
            order.append(k)
        seen[k].append(mr)

    rows = []
    for k in order:
        mrs   = seen[k]
        first = mrs[0]
        pi    = first.pdf_item
        ni    = first.netto_item

        total_qty = sum(mr.pdf_item.quantity for mr in mrs)
        secs: List[str] = []
        for mr in mrs:
            s = mr.pdf_item.section if mr.pdf_item else ""
            if s and s not in secs:
                secs.append(s)

        ib_qty     = first.ib_qty
        qty_diff   = round(total_qty - ib_qty, 4) if ib_qty is not None else None
        price_diff = first.price_diff
        pct_diff   = (
            f"{price_diff / ni.netto_price * 100:+.1f}%"
            if price_diff is not None and ni and ni.netto_price else ""
        )

        has_qty   = qty_diff   is not None and qty_diff   != 0
        has_price = price_diff is not None and abs(price_diff) > 0.01
        if   first.pdf_match_method == "unmatched": status = "unmatched"
        elif has_qty:                               status = "qty_diff"
        elif has_price:                             status = "price_diff"
        else:                                       status = "ok"

        rows.append({
            "_status":          status,
            "Art.nr.":          pi.art_nr,
            "Manual nr.":       ni.manual_nr        if ni else "",
            "Omschrijving":     pi.description,
            "Sectie":           " / ".join(secs),
            "PDF Aantal":       total_qty,
            "IB Aantal":        ib_qty              if ib_qty      is not None else None,
            "Aantal verschil":  qty_diff            if qty_diff    is not None else None,
            "PDF Prijs p.e.":   pi.unit_price,
            "NETTO Prijs p.e.": ni.netto_price      if ni          else None,
            "Prijs verschil €": price_diff          if price_diff  is not None else None,
            "Prijs verschil %": pct_diff,
            "Methode":          first.pdf_match_method,
            "Match %":          f"{first.confidence:.0%}" if first.confidence else "",
            "Pagina":           pi.page,
        })
    return pd.DataFrame(rows)


def _style(df: pd.DataFrame) -> "pd.io.formats.style.Styler":
    """Apply row background colors from _status column, then hide it."""
    status = df["_status"].reset_index(drop=True)
    display = df.drop(columns=["_status"])

    def row_bg(row):
        color = _STATUS_COLOR.get(status.iloc[row.name], "")
        return [f"background-color:{color}"] * len(row)

    money = lambda x: f"€ {x:,.2f}"   if isinstance(x, (int, float)) else ""
    delta = lambda x: f"{x:+,.2f}"    if isinstance(x, (int, float)) else ""
    num   = lambda x: f"{x:g}"        if isinstance(x, (int, float)) else ""
    sgn   = lambda x: f"{x:+g}"       if isinstance(x, (int, float)) else ""

    return (
        display.style
        .apply(row_bg, axis=1)
        .format({
            "PDF Prijs p.e.":   money,
            "NETTO Prijs p.e.": money,
            "Prijs verschil €": delta,
            "PDF Aantal":       num,
            "IB Aantal":        num,
            "Aantal verschil":  sgn,
        }, na_rep="")
    )


# ─── Excel Export ──────────────────────────────────────────────────────────────

def export_to_bytes(df: pd.DataFrame) -> bytes:
    F = {
        "header":   PatternFill("solid", fgColor="4472C4"),
        "ok":       PatternFill("solid", fgColor="C6EFCE"),
        "qty_diff": PatternFill("solid", fgColor="FFCC00"),
        "price_diff": PatternFill("solid", fgColor="FFF2CC"),
        "unmatched":  PatternFill("solid", fgColor="FFC7CE"),
    }
    FT_HDR = Font(bold=True, color="FFFFFF")
    COL_W  = {"Omschrijving": 55, "Sectie": 22, "Manual nr.": 14,
               "Art.nr.": 12, "Methode": 14, "Match %": 10}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vergelijking"

    cols = [c for c in df.columns if c != "_status"]
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = F["header"]
        c.font = FT_HDR
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(h, 13)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        fill = F.get(row["_status"])
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row[col])
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Streamlit App ─────────────────────────────────────────────────────────────

def main():
    st.markdown("""
    <div class="jumbo-hdr">
        <div style="font-size:42px;line-height:1">🟡</div>
        <div>
            <h1>Offerte Vergelijker</h1>
            <p>Van Keulen — vergelijk PDF offerte met NETTO prijslijst en IB Budget</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Sidebar ────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.header("📂 Bestanden uploaden")
        pdf_file    = st.file_uploader("PDF Offerte",      type=["pdf"])
        netto_file  = st.file_uploader("NETTO Prijslijst", type=["xlsx"])
        budget_file = st.file_uploader("Budget (IB)",      type=["xlsm", "xlsx"])

        st.divider()
        if not HAS_RAPIDFUZZ:
            st.warning("rapidfuzz niet gevonden – fuzzy matching uitgeschakeld.\n\n`pip install rapidfuzz`")

        run = st.button("▶  Analyseren", type="primary", use_container_width=True)

        st.divider()
        st.markdown("""
        **Legenda**
        <div class="legend-row" style="flex-direction:column;gap:6px">
          <span><span class="legend-chip" style="background:#C6EFCE">OK</span> Exact / manual match</span>
          <span><span class="legend-chip" style="background:#FFCC00">⚠</span> Aantal verschil</span>
          <span><span class="legend-chip" style="background:#FFF2CC">€</span> Prijs verschil</span>
          <span><span class="legend-chip" style="background:#FFC7CE">✗</span> Niet gematch</span>
        </div>
        """, unsafe_allow_html=True)

    # ── Session state ──────────────────────────────────────────────────────────
    if "df" not in st.session_state:
        st.session_state.df      = None
        st.session_state.results = []

    # ── Run ────────────────────────────────────────────────────────────────────
    if run:
        if not pdf_file or not netto_file:
            st.error("Selecteer minimaal een **PDF offerte** en **NETTO prijslijst**.")
        else:
            _run_analysis(pdf_file, netto_file, budget_file)

    # ── Display ────────────────────────────────────────────────────────────────
    if st.session_state.df is not None:
        _show_results(st.session_state.df)
    else:
        st.info("Upload bestanden via de zijbalk en klik op **Analyseren** om te beginnen.")


def _run_analysis(pdf_file, netto_file, budget_file):
    bar = st.progress(0, "PDF lezen…")

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_file.read())
        pdf_path = tmp.name
    try:
        pdf_items = parse_pdf(pdf_path)
    finally:
        os.unlink(pdf_path)

    bar.progress(30, f"PDF: {len(pdf_items)} items. NETTO lezen…")
    netto_items = read_netto(io.BytesIO(netto_file.read()))

    bar.progress(55, f"NETTO: {len(netto_items)} items. IB Budget lezen…")
    ib_items: List[IBItem] = []
    if budget_file:
        ib_items = read_ib_items(io.BytesIO(budget_file.read()))

    bar.progress(75, f"Matchen ({len(pdf_items)} PDF × {len(netto_items)} NETTO)…")
    results = build_matches(pdf_items, netto_items, ib_items)
    df = results_to_df(results)

    st.session_state.df      = df
    st.session_state.results = results
    bar.progress(100, "Klaar!")
    bar.empty()
    st.rerun()


def _show_results(df: pd.DataFrame):
    # ── Metrics ────────────────────────────────────────────────────────────────
    total     = len(df)
    exact     = int((df["Methode"] == "exact_artnr").sum())
    manual    = int((df["Methode"] == "manual_code").sum())
    fuzzy     = int((df["Methode"] == "fuzzy").sum())
    unmatched = int((df["Methode"] == "unmatched").sum())
    qty_afw   = int((df["_status"] == "qty_diff").sum())
    prc_col   = pd.to_numeric(df["Prijs verschil €"], errors="coerce")
    prc_afw   = int((prc_col.notna() & (prc_col.abs() > 0.01)).sum())

    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    c1.metric("Unieke art.nr.", total)
    c2.metric("Exact",         exact,     help="Exact artikel-nr. match")
    c3.metric("Manual code",   manual,    help="Match via Manual-header in PDF")
    c4.metric("Fuzzy",         fuzzy,     help="Fuzzy beschrijvingsMatch")
    c5.metric("Niet gematch",  unmatched)
    c6.metric("Aantal ≠",      qty_afw,   help="Aantal verschil PDF vs IB Budget")
    c7.metric("Prijs ≠",       prc_afw,   help="Prijs verschil PDF vs NETTO")

    # ── Tabs ───────────────────────────────────────────────────────────────────
    tab_all, tab_rev, tab_exp = st.tabs(["📋 Alle resultaten", "⚠️ Te controleren", "💾 Export"])

    with tab_all:
        # Filters
        fc1, fc2, fc3 = st.columns([2, 2, 4])
        with fc1:
            meth_f = st.multiselect(
                "Methode", df["Methode"].unique().tolist(), key="filt_meth",
            )
        with fc2:
            secs = sorted(s for s in df["Sectie"].unique() if s)
            sec_f = st.multiselect("Sectie", secs, key="filt_sec")
        with fc3:
            search = st.text_input("Zoeken (omschrijving / art.nr.)", key="filt_search")

        fdf = df.copy()
        if meth_f:  fdf = fdf[fdf["Methode"].isin(meth_f)]
        if sec_f:   fdf = fdf[fdf["Sectie"].isin(sec_f)]
        if search:
            q = search.lower()
            fdf = fdf[
                fdf["Omschrijving"].str.lower().str.contains(q, na=False) |
                fdf["Art.nr."].astype(str).str.lower().str.contains(q, na=False)
            ]

        st.caption(f"{len(fdf)} van {len(df)} rijen getoond")
        st.dataframe(_style(fdf.reset_index(drop=True)), use_container_width=True, height=540)

    with tab_rev:
        rdf = df[df["Methode"].isin(["fuzzy", "unmatched"])].reset_index(drop=True)
        if rdf.empty:
            st.success("Geen items te controleren — alles exact of via manual code gematch!")
        else:
            st.caption(f"{len(rdf)} items vereisen handmatige controle")
            st.dataframe(_style(rdf), use_container_width=True, height=500)

    with tab_exp:
        st.subheader("Resultaten downloaden")
        ec1, ec2 = st.columns(2)
        with ec1:
            st.download_button(
                "📥  Download Excel (met kleurcodering)",
                data=export_to_bytes(df),
                file_name="Vergelijking_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with ec2:
            csv = df.drop(columns=["_status"]).to_csv(index=False, sep=";").encode("utf-8-sig")
            st.download_button(
                "📥  Download CSV",
                data=csv,
                file_name="Vergelijking_output.csv",
                mime="text/csv",
                use_container_width=True,
            )
        st.info("De Excel bevat dezelfde kleurcodering als de tabelweergave hierboven.")


if __name__ == "__main__":
    main()
