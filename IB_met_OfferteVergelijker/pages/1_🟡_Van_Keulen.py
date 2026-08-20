#!/usr/bin/env python3
"""
Offerte Vergelijker – Van Keulen section.
Run via the app entrypoint:  streamlit run IB_met_OfferteVergelijker/Home.py
"""

import io
import os
import tempfile
from typing import Dict, List

import pandas as pd
import streamlit as st

from common import configure_page, inject_base_style, jumbo_header, back_to_overview, VAN_KEULEN_ICON

try:
    from matching_van_keulen import (
        IBItem,
        HAS_RAPIDFUZZ,
        apply_art_nr_mapping,
        build_matches,
        load_art_nr_mapping,
        parse_pdf,
        read_budget_summary_rows,
        read_ib_items,
        read_ib_row_overrides,
        read_netto,
        results_to_df,
    )
except ImportError as e:
    st.error(f"Ontbrekende dependency — `pip install -r requirements.txt` ({e})")
    st.stop()

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    st.error("openpyxl niet gevonden — `pip install openpyxl`")
    st.stop()

# ── Page config ────────────────────────────────────────────────────────────────
configure_page("Offerte Vergelijker – Van Keulen", icon=VAN_KEULEN_ICON)
inject_base_style()


def _style(df: pd.DataFrame) -> "pd.io.formats.style.Styler":
    """Cell-level highlights only: orange on Aantal verschil, red on Art.nr. for unmatched."""
    status  = df["_status"].reset_index(drop=True)
    display = df.drop(columns=["_status", "Methode"])

    def qty_cell(col):
        return [
            "background-color:#FFCC00" if status.iloc[i] == "qty_diff" else ""
            for i in range(len(col))
        ]

    def unmatched_cell(col):
        return [
            "background-color:#FFC7CE; font-weight:600" if status.iloc[i] == "unmatched" else ""
            for i in range(len(col))
        ]

    money = lambda x: f"€ {x:,.2f}" if isinstance(x, (int, float)) else ""
    num   = lambda x: f"{x:g}"      if isinstance(x, (int, float)) else ""
    sgn   = lambda x: f"{x:+g}"     if isinstance(x, (int, float)) else ""

    return (
        display.style
        .apply(qty_cell,       subset=["Aantal verschil"])
        .apply(unmatched_cell, subset=["Art.nr."])
        .format({
            "PDF Prijs p.e.":   money,
            "PDF Totaal":       money,
            "NETTO Prijs p.e.": money,
            "PDF Aantal":       num,
            "IB Aantal":        num,
            "Aantal verschil":  sgn,
        }, na_rep="")
    )


# ─── Excel Export ──────────────────────────────────────────────────────────────

def export_to_bytes(df: pd.DataFrame) -> bytes:
    F = {
        "header":   PatternFill("solid", fgColor="4472C4"),
        "ok":       PatternFill("solid", fgColor="C6EFCE"),
        "qty_diff": PatternFill("solid", fgColor="FFCC00"),
        "unmatched":PatternFill("solid", fgColor="FFC7CE"),
    }
    FT_HDR = Font(bold=True, color="FFFFFF")
    COL_W  = {"Omschrijving": 55, "Sectie": 22, "Manual nr.": 14,
               "Art.nr.": 12, "Methode": 14, "Match %": 10}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vergelijking"

    cols = [c for c in df.columns if c not in ("_status", "Methode")]
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = F["header"]
        c.font = FT_HDR
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(h, 13)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        fill = F.get(row["_status"])
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row[col])
            if fill:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Streamlit App ─────────────────────────────────────────────────────────────

def main():
    back_to_overview()
    jumbo_header(VAN_KEULEN_ICON, "Offerte Vergelijker", "Van Keulen — vergelijk PDF offerte met NETTO prijslijst en IB Budget")

    # ── Sidebar ────────────────────────────────────────────────────────────────
    with st.sidebar:
        st.header("📂 Bestanden uploaden")
        pdf_file    = st.file_uploader("PDF Offerte",      type=["pdf"])
        netto_file  = st.file_uploader("NETTO Prijslijst", type=["xlsx"])
        budget_file = st.file_uploader("Budget (IB)",      type=["xlsm", "xlsx"])

        st.divider()
        if not HAS_RAPIDFUZZ:
            st.warning("rapidfuzz niet gevonden – fuzzy matching uitgeschakeld.\n\n`pip install rapidfuzz`")

        run = st.button("▶  Analyseren", type="primary", use_container_width=True)

        st.divider()
        st.markdown("""
        **Legenda**
        <div class="legend-row" style="flex-direction:column;gap:6px">
          <span><span class="legend-chip" style="background:#C6EFCE">OK</span> Exact / manual match</span>
          <span><span class="legend-chip" style="background:#FFCC00">⚠</span> Aantal verschil</span>
          <span><span class="legend-chip" style="background:#FFC7CE">✗</span> Niet gematch</span>
          <span>↑ / ↓ Prijs verschil (alleen bij exact art.nr. match)</span>
        </div>
        """, unsafe_allow_html=True)

    # ── Session state ──────────────────────────────────────────────────────────
    if "df" not in st.session_state:
        st.session_state.df               = None
        st.session_state.results          = []
        st.session_state.budget_summary   = []
        st.session_state.lamellen_excluded = []
        st.session_state.raw_pdf_total    = 0.0
        st.session_state.raw_pdf_count    = 0
        st.session_state.mapping_filled   = 0
        st.session_state.mapping_size     = 0

    # ── Run ────────────────────────────────────────────────────────────────────
    if run:
        if not pdf_file or not netto_file:
            st.error("Selecteer minimaal een **PDF offerte** en **NETTO prijslijst**.")
        else:
            _run_analysis(pdf_file, netto_file, budget_file)

    # ── Display ────────────────────────────────────────────────────────────────
    if st.session_state.df is not None:
        if st.session_state.mapping_filled:
            st.caption(
                f"ℹ️ {st.session_state.mapping_filled} IB-regel(s) automatisch gekoppeld via het "
                f"bijgehouden mapping-bestand (van {st.session_state.mapping_size} bekende omschrijvingen) "
                "— kolom BA had voor deze regels geen Art.nr."
            )
        _show_results(st.session_state.df, st.session_state.budget_summary, st.session_state.lamellen_excluded)
    else:
        st.info("Upload bestanden via de zijbalk en klik op **Analyseren** om te beginnen.")


def _run_analysis(pdf_file, netto_file, budget_file):
    bar = st.progress(0, "PDF lezen…")

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_file.read())
        pdf_path = tmp.name
    try:
        pdf_items = parse_pdf(pdf_path)
    finally:
        os.unlink(pdf_path)

    bar.progress(30, f"PDF: {len(pdf_items)} items. NETTO lezen…")
    netto_items = read_netto(io.BytesIO(netto_file.read()))

    bar.progress(55, f"NETTO: {len(netto_items)} items. IB Budget lezen…")
    ib_items: List[IBItem] = []
    budget_summary: List[dict] = []
    lamellen_excluded: List[dict] = []
    ib_row_override: Dict[str, IBItem] = {}
    art_nr_mapping = load_art_nr_mapping()
    mapping_filled = 0
    if budget_file:
        budget_bytes = budget_file.read()
        ib_items       = read_ib_items(io.BytesIO(budget_bytes))
        mapping_filled = apply_art_nr_mapping(ib_items, art_nr_mapping, {p.art_nr for p in pdf_items})
        budget_summary, lamellen_excluded = read_budget_summary_rows(io.BytesIO(budget_bytes), 1986, 2763)
        ib_row_override = read_ib_row_overrides(io.BytesIO(budget_bytes))

    bar.progress(75, f"Matchen ({len(pdf_items)} PDF × {len(netto_items)} NETTO)…")
    results = build_matches(pdf_items, netto_items, ib_items, ib_row_override=ib_row_override)
    df = results_to_df(results)

    st.session_state.df                = df
    st.session_state.results           = results
    st.session_state.budget_summary    = budget_summary
    st.session_state.lamellen_excluded = lamellen_excluded
    st.session_state.raw_pdf_total     = sum(item.total for item in pdf_items)
    st.session_state.raw_pdf_count     = len(pdf_items)
    st.session_state.mapping_filled    = mapping_filled
    st.session_state.mapping_size      = len(art_nr_mapping)
    bar.progress(100, "Klaar!")
    bar.empty()
    st.rerun()


def _show_results(df: pd.DataFrame, budget_summary: List[dict] = None, lamellen_excluded: List[dict] = None):
    if lamellen_excluded:
        rijen = ", ".join(str(r["rij"]) for r in lamellen_excluded)
        st.error(
            f"⚠️ Lamellen Plafond will not be calculated with Van Keulen — "
            f"dit item hoort bij een andere leverancier (IB budget rij {rijen}). "
            f"Deze rij(en) zijn uitgesloten van het IB totaal."
        )

    # ── Metrics ────────────────────────────────────────────────────────────────
    total     = len(df)
    exact     = int((df["Methode"] == "exact_artnr").sum())
    manual    = int((df["Methode"] == "manual_code").sum())
    fuzzy     = int((df["Methode"] == "fuzzy").sum())
    unmatched = int((df["Methode"] == "unmatched").sum())
    qty_afw = int((df["_status"] == "qty_diff").sum())
    prc_afw = int((df["Prijs verschil"] != "").sum())

    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    c1.metric("Unieke art.nr.", total)
    c2.metric("Exact",         exact,     help="Exact artikel-nr. match")
    c3.metric("Manual code",   manual,    help="Match via Manual-header in PDF")
    c4.metric("Fuzzy",         fuzzy,     help="Fuzzy beschrijvingsMatch")
    c5.metric("Niet gematch",  unmatched)
    c6.metric("Aantal ≠",      qty_afw,   help="Aantal verschil PDF vs IB Budget")
    c7.metric("Prijs ≠",       prc_afw,   help="Prijs verschil PDF vs NETTO")

    # ── Tabs ───────────────────────────────────────────────────────────────────
    # Samenvatting first — the tab a user should land on right after
    # Analyseren, per project convention (see CLAUDE.md).
    tab_sam, tab_all, tab_rev, tab_exp = st.tabs(["📊 Samenvatting", "📋 Alle resultaten", "⚠️ Te controleren", "💾 Export"])

    with tab_all:
        # Filters
        fc1, fc2, fc3 = st.columns([2, 2, 4])
        with fc1:
            meth_f = st.multiselect(
                "Methode", df["Methode"].unique().tolist(), key="filt_meth",
            )
        with fc2:
            secs = sorted(s for s in df["Sectie"].unique() if s)
            sec_f = st.multiselect("Sectie", secs, key="filt_sec")
        with fc3:
            search = st.text_input("Zoeken (omschrijving / art.nr.)", key="filt_search")

        fdf = df.copy()
        if meth_f:  fdf = fdf[fdf["Methode"].isin(meth_f)]
        if sec_f:   fdf = fdf[fdf["Sectie"].isin(sec_f)]
        if search:
            q = search.lower()
            fdf = fdf[
                fdf["Omschrijving"].str.lower().str.contains(q, na=False) |
                fdf["Art.nr."].astype(str).str.lower().str.contains(q, na=False)
            ]

        st.caption(f"{len(fdf)} van {len(df)} rijen getoond")
        st.dataframe(_style(fdf.reset_index(drop=True)), use_container_width=True, height=540)

    with tab_rev:
        rdf = df[df["_status"] == "qty_diff"].reset_index(drop=True)
        if rdf.empty:
            st.success("Geen items te controleren — geen aantalverschillen gevonden!")
        else:
            st.caption(f"{len(rdf)} items met aantalverschil vereisen handmatige controle")
            st.dataframe(_style(rdf), use_container_width=True, height=500)

    with tab_sam:
        pdf_total      = float(pd.to_numeric(df["PDF Totaal"], errors="coerce").sum())
        raw_pdf_total  = st.session_state.raw_pdf_total
        raw_pdf_count  = st.session_state.raw_pdf_count

        if not budget_summary:
            st.info("Upload een Budget (IB) bestand om de budgetrijen te tonen.")
            budget_total = 0.0
        else:
            budget_total = sum(r["totaal"] for r in budget_summary)

        diff      = pdf_total - budget_total
        diff_pct  = diff / budget_total * 100 if budget_total else 0

        # ── Top KPIs ──────────────────────────────────────────────────────────
        k1, k2, k3 = st.columns(3)
        k1.metric("Totaal PDF offerte",  f"€ {pdf_total:,.2f}")
        k2.metric("Totaal Budget (IB)",  f"€ {budget_total:,.2f}")
        diff_label = f"{'▲' if diff > 0 else '▼'} € {abs(diff):,.2f}  ({diff_pct:+.1f}%)"
        k3.metric("Verschil", diff_label, delta_color="inverse" if diff > 0 else "normal")

        # ── Parser diagnostic ─────────────────────────────────────────────────
        with st.expander("🔍 Parser diagnose (klik om te openen)"):
            d1, d2, d3 = st.columns(3)
            d1.metric("Regels geparseerd uit PDF", raw_pdf_count)
            d2.metric("Raw PDF totaal (voor groepering)", f"€ {raw_pdf_total:,.2f}")
            d3.metric("Na groepering op art.nr.", f"€ {pdf_total:,.2f}")
            if abs(raw_pdf_total - pdf_total) > 0.01:
                st.warning(f"Verschil door groepering: € {raw_pdf_total - pdf_total:,.2f} — mogelijk duplicaat art.nr. met verschillende prijzen.")
            missing = 161599 - raw_pdf_total
            if abs(missing) > 0.01:
                st.error(f"Ontbrekend t.o.v. verwacht totaal (€ 161.599): € {missing:,.2f} — waarschijnlijk regels die de parser mist.")

        st.divider()

        # ── Budget rows breakdown ──────────────────────────────────────────────
        if budget_summary:
            st.subheader("Budget rijen 1986 – 2763 (IB)")
            bdf = pd.DataFrame(budget_summary).rename(columns={
                "rij": "Rij", "nummer": "Nummer", "artikelnaam": "Artikelnaam",
                "eenheid": "Eenheid", "aantal": "Aantal",
                "prijs": "Prijs p.e.", "totaal": "Totaal",
            })
            st.caption(f"{len(bdf)} rijen met aantal ≥ 1  |  totaal: € {budget_total:,.2f}")
            st.dataframe(
                bdf.style.format({
                    "Prijs p.e.": lambda x: f"€ {x:,.2f}",
                    "Totaal":     lambda x: f"€ {x:,.2f}",
                    "Aantal":     lambda x: f"{x:g}",
                }, na_rep=""),
                use_container_width=True,
                hide_index=True,
                height=350,
            )

        st.divider()

        # ── PDF totaal per sectie ──────────────────────────────────────────────
        st.subheader("PDF totaal per sectie")
        sec_df = (
            df.assign(**{"PDF Totaal num": pd.to_numeric(df["PDF Totaal"], errors="coerce")})
            .groupby("Sectie", dropna=False)["PDF Totaal num"]
            .sum()
            .reset_index()
            .rename(columns={"Sectie": "Sectie", "PDF Totaal num": "Totaal"})
            .sort_values("Totaal", ascending=False)
        )
        sec_df["Totaal"] = sec_df["Totaal"].apply(lambda x: f"€ {x:,.2f}")
        st.dataframe(sec_df, use_container_width=True, hide_index=True)

    with tab_exp:
        st.subheader("Resultaten downloaden")
        ec1, ec2 = st.columns(2)
        with ec1:
            st.download_button(
                "📥  Download Excel (met kleurcodering)",
                data=export_to_bytes(df),
                file_name="Vergelijking_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with ec2:
            csv = df.drop(columns=["_status"]).to_csv(index=False, sep=";").encode("utf-8-sig")
            st.download_button(
                "📥  Download CSV",
                data=csv,
                file_name="Vergelijking_output.csv",
                mime="text/csv",
                use_container_width=True,
            )
        st.info("De Excel bevat dezelfde kleurcodering als de tabelweergave hierboven.")


main()
