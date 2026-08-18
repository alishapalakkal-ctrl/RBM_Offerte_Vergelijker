#!/usr/bin/env python3
"""
Offerte Vergelijker – Koeling section.
Run via the app entrypoint:  streamlit run IB_met_OfferteVergelijker/Home.py

Two independent sub-flows, chosen up front:

  Koel Installatie
    Compares the Dutch koelinstallatie-offerte (bidsheet, e.g. Frimex
    'Offertetemplate KT') against the IB macro-werkboek ('2024 Installatie
    <leverancier>' sheet). No shared article number exists between the two
    by default — matching uses a manually filled-in 'Art.Nr.Jumbo' column
    in the offerte (exact join) with a description-fuzzy fallback for the
    rest, then shows aantal- and prijsverschillen per matched item.

  Koel Meubel
    Unlike Koel Installatie, the matching itself happens in Dynamo (Revit
    model position numbers matched to elements via bounding-box
    containment, see 02_Positienummer Match.dyn). This just visualizes
    what that graph exports:
      - a single combined Excel file (3 tabs: Vergelijking, Posities niet
        in offerte, Accessoires) from the 'Compare & Export Koeling' node
      - the colored layout PDF (green/red per position) from the graph's
        'Export Koeling View to PDF' node
"""

import io

import altair as alt
import openpyxl
import pandas as pd
import streamlit as st
from streamlit_pdf_viewer import pdf_viewer

from common import configure_page, inject_base_style, jumbo_header, back_to_overview
from matching_koeling import (
    HAS_RAPIDFUZZ,
    apply_art_nr_mapping,
    build_budget_link,
    build_buffetten_link,
    build_categorie_totalen,
    build_installatie_df,
    load_art_nr_mapping,
    load_budget_df,
    load_categorie_budget_codes,
    match_koeling_installatie,
    parse_ib,
    parse_nrfp,
    parse_offerte,
    _pick_koeling_sheets,
)

configure_page("Offerte Vergelijker – Koeling", icon="🧊")
inject_base_style()


#: Koelcellen, vriescellen and celaccessoires are one physical scope from
#: the koelinstallateur's perspective (the cellen plus the accessories that
#: go with them) even though the offerte/Budget book them as three separate
#: categories — grouped together for the verschil-per-categorie pie chart
#: so it reads as one slice instead of three slivers.
_PIE_CATEGORIE_GROEP = {
    "Categorie koelcellen": "Koel-/vriescellen (incl. celaccessoires)",
    "Categorie vriescellen": "Koel-/vriescellen (incl. celaccessoires)",
    "Categorie celaccessoires": "Koel-/vriescellen (incl. celaccessoires)",
}


def _pie_categorie_label(categorie: str) -> str:
    grouped = _PIE_CATEGORIE_GROEP.get(categorie, categorie)
    return grouped[len("Categorie "):] if grouped.startswith("Categorie ") else grouped


# ═══════════════════════════════════════════════════════════════════════════
# Koel Installatie — offerte (bidsheet) vs IB matching
# ═══════════════════════════════════════════════════════════════════════════

def _style_installatie(df: pd.DataFrame):
    status = df["_status"].reset_index(drop=True)
    hidden = [c for c in df.columns if c.startswith("_")] + ["Match", "Prijs verschil (€)", "Art.Nr.Jumbo"]
    display = df.drop(columns=hidden).rename(columns={"Prijs verschil (%)": "Prijs verschil"})

    def _status_cell(target_status: str, css: str):
        def styler(col):
            return [css if status.iloc[i] == target_status else "" for i in range(len(col))]
        return styler

    money = lambda x: f"€ {x:,.2f}" if isinstance(x, (int, float)) else ""
    num = lambda x: f"{x:g}" if isinstance(x, (int, float)) else ""

    def prijs_arrow(x):
        if not isinstance(x, (int, float)):
            return ""
        if x > 0.05:
            return f"▲ {x:.1f}%"
        if x < -0.05:
            return f"▼ {abs(x):.1f}%"
        return ""

    return (
        display.style
        .apply(_status_cell("unmatched", "background-color:#FFC7CE; font-weight:600"), subset=["IB omschrijving"])
        .apply(_status_cell("aantal_afwijking", "background-color:#FFCC00"),
               subset=["Offerte Aantal", "IB Aantal", "Aantal verschil"])
        .format({
            "Offerte Prijs p.e.": money, "IB Prijs p.e.": money,
            "Prijs verschil": prijs_arrow,
            "Offerte Aantal": num, "IB Aantal": num, "Aantal verschil": num,
        }, na_rep="")
    )


def _run_koel_installatie():
    jumbo_header("🧊", "Offerte Vergelijker", "Koeling — Koel Installatie")

    with st.sidebar:
        st.header("📂 Bestanden uploaden")
        ib_file = st.file_uploader(
            "IB macro-werkboek (.xlsm)", type=["xlsm"],
            help="Het macro-werkboek met 'IB' in de bestandsnaam, bevat de sheet '2024 Installatie <leverancier>'.",
        )
        bid_file = st.file_uploader(
            "Bidsheet offerte (.xlsx)", type=["xlsx"],
            help="De offerte-Excel met 'bidsheet' in de bestandsnaam (bijv. 'Offertetemplate KT').",
        )
        st.caption(
            "Matching gebeurt via een Art.Nr.Jumbo per omschrijving — uit de offerte zelf als die "
            "kolom is ingevuld, anders uit een intern bijgehouden mapping-bestand — met fuzzy "
            "tekst-matching als terugval voor de rest."
        )
        run = st.button("▶ Analyseren", type="primary", use_container_width=True)
        if not HAS_RAPIDFUZZ:
            st.warning("rapidfuzz niet gevonden – fuzzy matching uitgeschakeld.")

    if "koel_inst_df" not in st.session_state:
        st.session_state.koel_inst_df = None
        st.session_state.koel_inst_mapping_filled = 0
        st.session_state.koel_inst_mapping_size = 0
        st.session_state.koel_inst_nrfp_items = []
        st.session_state.koel_inst_categorie_totalen = {}

    if run:
        if not ib_file or not bid_file:
            st.error("Upload zowel het **IB macro-werkboek** als de **bidsheet offerte**.")
        else:
            _analyze_koel_installatie(ib_file, bid_file)

    if st.session_state.koel_inst_df is None:
        st.info("Upload beide bestanden via de zijbalk en klik op **Analyseren** om te beginnen.")
        return

    if st.session_state.koel_inst_mapping_filled:
        st.caption(
            f"ℹ️ {st.session_state.koel_inst_mapping_filled} regel(s) automatisch gekoppeld via het "
            f"bijgehouden mapping-bestand (van {st.session_state.koel_inst_mapping_size} bekende "
            "omschrijvingen) — de offerte zelf had geen Art.Nr.Jumbo voor deze regels."
        )

    _show_koel_installatie_results(
        st.session_state.koel_inst_df,
        st.session_state.koel_inst_nrfp_items,
        st.session_state.koel_inst_categorie_totalen,
    )


def _analyze_koel_installatie(ib_file, bid_file):
    ib_bytes = ib_file.read()
    bid_bytes = bid_file.read()

    # Only used to list sheet names for the selectboxes below (cheap —
    # doesn't require parsing individual sheets' contents). Actual row data
    # is read via pandas in parse_offerte/parse_ib/build_budget_link, which
    # is dramatically faster than openpyxl's read_only=True random .cell()
    # access on this large, complex workbook.
    wb_ib = openpyxl.load_workbook(io.BytesIO(ib_bytes), read_only=True, data_only=True)
    wb_bid = openpyxl.load_workbook(io.BytesIO(bid_bytes), read_only=True, data_only=True)

    ib_sheets = _pick_koeling_sheets(wb_ib)
    if not ib_sheets:
        st.error("Geen sheet gevonden met 'Installatie' in de naam in het IB-bestand.")
        return
    bid_sheets = [s for s in wb_bid.sheetnames if "offertetemplate" in s.lower()] or wb_bid.sheetnames

    c1, c2 = st.columns(2)
    ib_sheet = c1.selectbox("IB sheet", ib_sheets, index=0, key="koel_inst_ib_sheet")
    bid_sheet = c2.selectbox("Offerte sheet", bid_sheets, index=0, key="koel_inst_bid_sheet")

    with st.spinner("Analyseren…"):
        offerte = parse_offerte(bid_bytes, bid_sheet)
        art_nr_mapping = load_art_nr_mapping("frimex")
        mapping_filled = apply_art_nr_mapping(offerte, art_nr_mapping)
        ib = parse_ib(ib_bytes, ib_sheet)
        matches, _used_ib = match_koeling_installatie(offerte, ib)
        bdf = load_budget_df(ib_bytes)
        budget_link = build_budget_link(bdf, ib)
        buffetten_link = build_buffetten_link(bdf, offerte)
        df = build_installatie_df(offerte, ib, matches, budget_link, buffetten_link)
        nrfp_items = parse_nrfp(bid_bytes)
        categorie_mapping = load_categorie_budget_codes()
        categorie_totalen = build_categorie_totalen(bdf, categorie_mapping)

    st.session_state.koel_inst_df = df
    st.session_state.koel_inst_mapping_filled = mapping_filled
    st.session_state.koel_inst_mapping_size = len(art_nr_mapping)
    st.session_state.koel_inst_nrfp_items = nrfp_items
    st.session_state.koel_inst_categorie_totalen = categorie_totalen
    st.rerun()


def _show_koel_installatie_results(df: pd.DataFrame, nrfp_items=None, categorie_totalen=None):
    nrfp_items = nrfp_items or []
    categorie_totalen = categorie_totalen or {}
    # Only show offerte lines that were actually selected for this project
    # (Offerte Aantal filled in) — most rows are unused catalog alternatives
    # (e.g. Boosterset 1-5 when only one size was chosen) with no aantal.
    df = df[df["Offerte Aantal"].notna()].reset_index(drop=True)

    total = len(df)
    handmatig = int((df["Match"] == "handmatig").sum())
    fuzzy = int((df["Match"] == "fuzzy").sum())
    unmatched = int((df["_status"] == "unmatched").sum())
    prijs_afw = int((df["_status"] == "prijs_afwijking").sum())
    aantal_afw = int((df["_status"] == "aantal_afwijking").sum())

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Offerte regels", total)
    c2.metric("Handmatig", handmatig, help="Gematcht via Art.Nr.Jumbo kolom")
    c3.metric("Fuzzy", fuzzy, help="Gematcht via tekst-vergelijking")
    c4.metric("Niet gematcht", unmatched)
    c5.metric("Prijs ≠", prijs_afw)
    c6.metric("Aantal ≠", aantal_afw)

    # Samenvatting first — the tab a user should land on right after
    # Analyseren, per project convention (see CLAUDE.md).
    tab_sam, tab_all, tab_fuzzy, tab_exp = st.tabs(
        ["📊 Samenvatting", "📋 Alle resultaten", "🔎 Fuzzy matches (controleren)", "💾 Export"]
    )

    with tab_all:
        st.caption(
            f"{total} offerte-regels — rood = geen IB-match, geel = aantalverschil > 10%, "
            "▲/▼ in Prijs verschil = prijsafwijking"
        )
        st.dataframe(_style_installatie(df), use_container_width=True, height=600)

    with tab_fuzzy:
        fdf = df[df["Match"] == "fuzzy"].reset_index(drop=True)
        if fdf.empty:
            st.success("Alle matches zijn handmatig bevestigd via Art.Nr.Jumbo.")
        else:
            st.caption(
                f"{len(fdf)} regels zijn via fuzzy tekst-matching gekoppeld — controleer deze en vul zo nodig "
                "handmatig een Art.Nr.Jumbo in op de offerte voor een zekere match."
            )
            st.dataframe(_style_installatie(fdf), use_container_width=True, height=500)

    with tab_sam:
        offerte_totaal = float(df["_offerte_totaal"].sum(skipna=True))

        # cat_df is built once, up front, and reused for the top "Totaal IB"
        # KPI, the categorie table and the pie chart below — its "Totaal IB"
        # already prefers the Budget subheading's own Totaal (via
        # categorie_totalen) over the per-row rollup, which is the more
        # realistic figure: the per-row rollup reads 0 for a variant the
        # offerte quoted but IB ended up not selecting.
        cat_df = (
            df.groupby("Categorie", dropna=False)[["_offerte_totaal", "_ib_totaal"]]
            .sum(min_count=1)
            .reset_index()
            .rename(columns={"_offerte_totaal": "Totaal Offerte", "_ib_totaal": "Totaal IB"})
        )
        if categorie_totalen:
            cat_df["Totaal IB"] = cat_df.apply(
                lambda r: categorie_totalen.get(r["Categorie"], r["Totaal IB"]), axis=1
            )
        cat_df["Verschil"] = cat_df["Totaal Offerte"] - cat_df["Totaal IB"]
        cat_df = cat_df.sort_values("Totaal Offerte", ascending=False, na_position="last")

        ib_totaal = float(cat_df["Totaal IB"].sum(skipna=True))
        diff = offerte_totaal - ib_totaal
        diff_pct = diff / ib_totaal * 100 if ib_totaal else 0

        k1, k2, k3 = st.columns(3)
        k1.metric("Totaal Offerte", f"€ {offerte_totaal:,.2f}",
                   help="(Prijs materiaal + Arbeidskosten) × Offerte Aantal, alleen regels met een ingevuld aantal")
        k2.metric("Totaal IB", f"€ {ib_totaal:,.2f}",
                   help="Som van Totaal per categorie hieronder — per categorie, waar bekend, de Budget "
                        "subtotaal-regel zelf, anders IB Prijs p.e. × Offerte Aantal per regel")
        diff_label = f"{'▲' if diff > 0 else '▼'} € {abs(diff):,.2f}  ({diff_pct:+.1f}%)"
        k3.metric("Verschil", diff_label, delta_color="inverse" if diff > 0 else "normal")

        st.divider()

        nrfp_totaal = sum(it.totaal for it in nrfp_items)
        st.subheader("NRFP — posities niet in de RFP")
        st.caption(
            "Extra/wijzigingsposities buiten de standaard tender (bijv. huur container, brandsturing) — "
            "de offerte's eigen Totaalprijs telt deze apart op naast de categorie-regels hierboven."
        )
        n1, n2 = st.columns(2)
        n1.metric("NRFP totaal", f"€ {nrfp_totaal:,.2f}")
        n2.metric("Totaal Offerte incl. NRFP", f"€ {offerte_totaal + nrfp_totaal:,.2f}")

        st.divider()

        st.subheader("Totaal per categorie")
        st.caption(
            "Totaal IB komt, waar bekend, rechtstreeks uit de bijbehorende subtotaal-regel op het "
            "Budget-tabblad (bijv. 'Gascooler') — dat blijft correct ook als IB een andere variant heeft "
            "gekozen dan de offerte, zie data/koeling_categorie_budget_codes.csv."
        )
        st.dataframe(
            cat_df.style.format({
                "Totaal Offerte": lambda x: f"€ {x:,.2f}" if pd.notna(x) else "",
                "Totaal IB": lambda x: f"€ {x:,.2f}" if pd.notna(x) else "",
                "Verschil": lambda x: f"€ {x:,.2f}" if pd.notna(x) else "",
            }),
            use_container_width=True, hide_index=True,
        )

        pie_df = cat_df.dropna(subset=["Verschil"]).copy()
        pie_df["Groep"] = pie_df["Categorie"].map(_pie_categorie_label)
        pie_df = pie_df.groupby("Groep", as_index=False)["Verschil"].sum()
        pie_df["Abs verschil"] = pie_df["Verschil"].abs()
        pie_df = pie_df[pie_df["Abs verschil"] > 0]

        if not pie_df.empty:
            c_pie, c_note = st.columns([1, 1])
            with c_pie:
                chart = (
                    alt.Chart(pie_df)
                    .mark_arc()
                    .encode(
                        theta=alt.Theta("Abs verschil:Q", title="Absoluut verschil (€)"),
                        color=alt.Color("Groep:N", title="Categorie"),
                        tooltip=[
                            alt.Tooltip("Groep:N", title="Categorie"),
                            alt.Tooltip("Verschil:Q", title="Verschil (€)", format=",.2f"),
                        ],
                    )
                )
                st.altair_chart(chart, use_container_width=True)
            with c_note:
                top = pie_df.sort_values("Abs verschil", ascending=False).iloc[0]
                top_share = top["Abs verschil"] / pie_df["Abs verschil"].sum() * 100
                st.markdown("**Analyse**")
                st.write(
                    f"Het grootste deel van het verschil zit in **{top['Groep']}** "
                    f"(€ {top['Verschil']:,.2f}, {top_share:.0f}% van het totale absolute verschil "
                    "tussen offerte en IB voor deze regels)."
                )
                st.caption(
                    "Een verschil in begroting hoeft geen fout te zijn — het kan ook ontstaan doordat IB en "
                    "de koelinstallateur op een andere manier rekenen: bijv. IB begroot per m² celvloer/"
                    "-wand of per project-eenheid, terwijl de installateur per stuk/component offreert, of "
                    "er is een andere verdeling tussen materiaal- en arbeidskosten aangehouden. Controleer bij "
                    "een groot verschil eerst of beide partijen dezelfde eenheid en scope hanteren voordat je "
                    "dit als een echte prijsafwijking behandelt."
                )

        st.divider()

        st.subheader("Match-methode overzicht")
        meth_df = df["Match"].value_counts().reset_index()
        meth_df.columns = ["Methode", "Aantal regels"]
        st.dataframe(meth_df, use_container_width=True, hide_index=True)

    with tab_exp:
        csv = df.drop(columns=[c for c in df.columns if c.startswith("_")]).to_csv(index=False, sep=";").encode("utf-8-sig")
        st.download_button(
            "📥 Download vergelijking (CSV)", data=csv,
            file_name="koel_installatie_vergelijking.csv", mime="text/csv",
        )


# ═══════════════════════════════════════════════════════════════════════════
# Koel Meubel — Dynamo-exported position match (existing flow)
# ═══════════════════════════════════════════════════════════════════════════

def _run_koel_meubel():
    if "koeling_leverancier" not in st.session_state:
        st.session_state.koeling_leverancier = None

    @st.dialog("Kies leverancier")
    def _choose_leverancier():
        st.write("Voor welke leverancier wil je offertes vergelijken?")
        c1, c2 = st.columns(2)
        if c1.button("Carrier", use_container_width=True):
            st.session_state.koeling_leverancier = "Carrier"
            st.rerun()
        if c2.button("Kaplanlaar", use_container_width=True):
            st.session_state.koeling_leverancier = "Kaplanlaar"
            st.rerun()

    if st.session_state.koeling_leverancier is None:
        jumbo_header("🧊", "Offerte Vergelijker", "Koeling — Koel Meubel — kies een leverancier om te starten")
        _choose_leverancier()
        st.stop()

    leverancier = st.session_state.koeling_leverancier
    jumbo_header("🧊", "Offerte Vergelijker", f"Koeling — Koel Meubel — {leverancier}")

    _, wc = st.columns([5, 1])
    with wc:
        if st.button("🔁 Wissel leverancier", use_container_width=True):
            st.session_state.koeling_leverancier = None
            st.rerun()

    with st.sidebar:
        st.header("📂 Bestanden uploaden")
        st.caption("Beide zijn exports van de Dynamo Positienummer Match graph.")
        excel_file = st.file_uploader(
            "Vergelijking (Excel, 3 tabbladen)",
            type=["xlsx"],
            help="Export van de 'Compare & Export Koeling' node — bevat Vergelijking, "
                 "Posities niet in offerte en Accessoires in één bestand.",
        )
        layout_pdf_file = st.file_uploader("Layout (PDF, met positienummers gekleurd)", type=["pdf"])

    sheets = pd.read_excel(excel_file, sheet_name=None) if excel_file else {}
    uploaded = {
        "Vergelijking": (excel_file, sheets.get("Vergelijking")),
        "Posities niet in offerte": (excel_file, sheets.get("Posities niet in offerte")),
        "Accessoires": (excel_file, sheets.get("Accessoires")),
    }

    if excel_file is None and layout_pdf_file is None:
        st.info(
            "Upload de vergelijkings-Excel en/of de layout-PDF via de zijbalk — beide zijn "
            "exports van de Dynamo Positienummer Match graph (zie 00_RBM_Offerte Vergelijker/Koeling)."
        )
        st.stop()

    tab_layout, tab_compare, tab_extra, tab_accessories = st.tabs(
        ["🗺️ Layout", "📋 Vergelijking", "➕ Posities niet in offerte", "🪟 Accessoires (Glass/Mirror)"]
    )

    with tab_layout:
        if layout_pdf_file is None:
            st.info("Upload de layout-PDF om deze tab te zien (export vanuit de 'Export Koeling View to PDF' Dynamo-node).")
        else:
            pdf_bytes = layout_pdf_file.getvalue()
            st.caption("Groen = code en lengte kloppen met de offerte, rood = afwijking.")
            pdf_viewer(pdf_bytes, width="100%", height=900)
            st.download_button(
                "📥 Download layout (PDF)",
                data=pdf_bytes,
                file_name=layout_pdf_file.name,
                mime="application/pdf",
            )

    with tab_compare:
        df = uploaded["Vergelijking"][1]
        if df is None:
            st.info("Upload de vergelijkings-Excel om deze tab te zien.")
        else:
            total = len(df)
            code_mismatch = int((df["code_match"] == False).sum())
            length_out = int((df["length_ok"] == False).sum())
            no_length = int(df["revit_length_mm"].isna().sum())

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Posities", total)
            c2.metric("Code afwijkend", code_mismatch)
            c3.metric("Lengte buiten tolerantie", length_out, help="Tolerantie: 80 mm")
            c4.metric("Geen lengte data", no_length)

            def _style_row(row):
                if row.get("code_match") is False or row.get("length_ok") is False:
                    return ["background-color:#FFC7CE"] * len(row)
                if pd.isna(row.get("revit_length_mm")):
                    return ["background-color:#FFCC00"] * len(row)
                return [""] * len(row)

            st.caption(f"{total} posities vergeleken — rood = code/lengte afwijkend, geel = geen lengte data")
            st.dataframe(df.style.apply(_style_row, axis=1), use_container_width=True, height=600)

            st.download_button(
                "📥 Download vergelijking (CSV)",
                data=df.to_csv(index=False).encode("utf-8-sig"),
                file_name="koeling_vergelijking.csv",
                mime="text/csv",
            )

    with tab_extra:
        df_extra = uploaded["Posities niet in offerte"][1]
        if df_extra is None:
            st.info("Upload de vergelijkings-Excel om deze tab te zien.")
        elif df_extra.empty:
            st.success("Geen extra posities gevonden — alle posities in het model staan ook in de offerte.")
        else:
            st.caption(
                f"{len(df_extra)} positie(s) gevonden in het Revit-model die niet in de offerte voorkomen "
                "(mogelijk door een andere leverancier gedekt, of ontbrekend in de offerte)."
            )
            st.dataframe(
                df_extra.rename(columns={
                    "pos": "Positienummer",
                    "manual_code": "Manual code",
                    "description": "Omschrijving",
                    "revit_length_mm": "Lengte (mm)",
                }),
                use_container_width=True,
                hide_index=True,
            )
            st.download_button(
                "📥 Download posities niet in offerte (CSV)",
                data=df_extra.to_csv(index=False).encode("utf-8-sig"),
                file_name="koeling_extra_posities.csv",
                mime="text/csv",
            )

    with tab_accessories:
        df_acc = uploaded["Accessoires"][1]
        if df_acc is None:
            st.info("Upload de vergelijkings-Excel om deze tab te zien.")
        else:
            total = len(df_acc)
            mismatches = int(((df_acc["glass_match"] == False) | (df_acc["mirror_match"] == False)).sum())

            c1, c2 = st.columns(2)
            c1.metric("Groepen (fysieke runs)", total)
            c2.metric("Afwijkend", mismatches)

            def _style_acc(row):
                if row.get("glass_match") is False or row.get("mirror_match") is False:
                    return ["background-color:#FFC7CE"] * len(row)
                if isinstance(row.get("source"), str) and "default" in row["source"]:
                    return ["background-color:#FFF2CC"] * len(row)
                return [""] * len(row)

            st.caption(
                f"{total} groepen — posities zijn gegroepeerd per fysieke run (opeenvolgende posities met "
                "dezelfde code, gesplitst op offerte-rijgrenzen). Rood = afwijkend, geel = default-aanname "
                "gebruikt (geen REF_Custom data op het element)."
            )
            st.dataframe(df_acc.style.apply(_style_acc, axis=1), use_container_width=True, height=600)

            st.download_button(
                "📥 Download accessoires (CSV)",
                data=df_acc.to_csv(index=False).encode("utf-8-sig"),
                file_name="koeling_accessoires.csv",
                mime="text/csv",
            )


# ═══════════════════════════════════════════════════════════════════════════
# Top-level: kies Koel Installatie of Koel Meubel
# ═══════════════════════════════════════════════════════════════════════════

if "koeling_kind" not in st.session_state:
    st.session_state.koeling_kind = None


@st.dialog("Kies type")
def _choose_kind():
    st.write("Wat wil je vergelijken?")
    c1, c2 = st.columns(2)
    if c1.button("🧯 Koel Installatie", use_container_width=True, help="IB macro-werkboek vs bidsheet offerte"):
        st.session_state.koeling_kind = "installatie"
        st.rerun()
    if c2.button("🧊 Koel Meubel", use_container_width=True, help="OC lijst / Dynamo Positienummer Match export"):
        st.session_state.koeling_kind = "meubel"
        st.rerun()


back_to_overview()

if st.session_state.koeling_kind is None:
    jumbo_header("🧊", "Offerte Vergelijker", "Koeling — kies een type om te starten")
    _choose_kind()
    st.stop()

_, wk = st.columns([5, 1])
with wk:
    if st.button("🔁 Wissel type", use_container_width=True):
        st.session_state.koeling_kind = None
        st.session_state.koeling_leverancier = None
        st.rerun()

if st.session_state.koeling_kind == "installatie":
    _run_koel_installatie()
else:
    _run_koel_meubel()
